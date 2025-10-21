"""Debug script to read Excel export using openpyxl."""
from openpyxl import load_workbook

# Load workbook
wb = load_workbook(r'D:\Automation\Скрины\ecademy_meta_data_2025-10-21.xlsx', data_only=True)

# Get Студенти sheet
ws = wb['Студенти']

# Get headers (first row)
headers = [cell.value for cell in ws[1]]
print(f"Total columns: {len(headers)}")
print("\nAll columns:")
for i, h in enumerate(headers, 1):
    print(f"  {i}. {h}")

# Get first data row (row 2)
print("\n" + "="*80)
print("FIRST ROW DATA:")
print("="*80)
for i, header in enumerate(headers, 1):
    value = ws.cell(row=2, column=i).value
    print(f"{header}: {value}")

# Find status columns
print("\n" + "="*80)
print("STATUS COLUMNS (first 3 data rows):")
print("="*80)
status_col_indices = [(i, h) for i, h in enumerate(headers, 1) if h and 'status_' in str(h).lower()]
print(f"Found {len(status_col_indices)} status columns")

if status_col_indices:
    for row_num in range(2, min(5, ws.max_row + 1)):  # rows 2-4
        print(f"\nRow {row_num-1}:")
        for col_idx, col_name in status_col_indices[:10]:  # first 10 status columns
            value = ws.cell(row=row_num, column=col_idx).value
            print(f"  {col_name}: {value}")

# Key metrics
print("\n" + "="*80)
print("KEY METRICS (first 3 data rows):")
print("="*80)
key_cols = ['campaign_name', 'leads_count', 'contacted', 'in_progress', 'purchased', 'status_13', 'status_32', 'status_4']
key_col_indices = []
for key_col in key_cols:
    try:
        idx = headers.index(key_col) + 1
        key_col_indices.append((idx, key_col))
    except ValueError:
        pass

for row_num in range(2, min(5, ws.max_row + 1)):
    print(f"\nRow {row_num-1}:")
    for col_idx, col_name in key_col_indices:
        value = ws.cell(row=row_num, column=col_idx).value
        print(f"  {col_name}: {value}")

print(f"\n\nTotal data rows: {ws.max_row - 1}")
