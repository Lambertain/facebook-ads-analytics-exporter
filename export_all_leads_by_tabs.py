"""
Експорт ВСІХ лідів з AlfaCRM в Excel таблицю з вкладками по статусах.

Формат:
- Кожна вкладка = 1 статус воронки
- Стовпці = поля ліда (ID, Ім'я, Телефони, Email, тощо)
- Рядки = окремі ліди
- Якщо воронок декілька - додаємо до назви статусу назву воронки
"""
import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
import requests
from collections import defaultdict
from typing import List, Dict, Any

# Загрузка переменных окружения
load_dotenv()

# Добавить app директорию в Python path
app_dir = Path(__file__).parent / "app"
sys.path.insert(0, str(app_dir))

from connectors.crm import alfacrm_auth_get_token


# Словарь переводов названий полей на украинский (с маркером С - системное, К - кастомное)
FIELD_TRANSLATIONS = {
    "id": "ID (С)",
    "name": "Ім'я (С)",
    "created_at": "Дата створення (С)",
    "updated_at": "Дата оновлення (С)",
    "lead_status_id": "ID статусу ліда (С)",
    "study_status_id": "ID статусу навчання (С)",
    "is_study": "Навчається (С)",
    "color": "Колір (С)",
    "phone": "Телефони (С)",
    "email": "Email адреси (С)",
    "addr": "Адреса (С)",
    "balance": "Баланс (С)",
    "balance_base": "Базовий баланс (С)",
    "balance_bonus": "Бонусний баланс (С)",
    "paid_count": "Кількість оплат (С)",
    "paid_lesson_count": "Кількість оплачених уроків (С)",
    "assigned_id": "ID відповідального (С)",
    "branch_ids": "ID філій (С)",
    "company_id": "ID компанії (С)",
    "b_date": "Дата народження (С)",
    "sex": "Стать (С)",
    "barcode": "Штрих-код (С)",
    "comment": "Коментар (С)",
    "note": "Note (С)",
    "legal_type": "Тип особи (С)",
    "legal_name": "Legal Name (С)",
    "pipeline_id": "Pipeline Id (С)",
    "web": "Web (С)",
    "teacher_ids": "Teacher Ids (С)",
    "last_attend_date": "Last Attend Date (С)",
    "next_lesson_date": "Next Lesson Date (С)",
    "paid_lesson_date": "Paid Lesson Date (С)",
    "paid_till": "Paid Till (С)",
    "e_date": "E Date (С)",
    "dob": "Dob (С)",
    "customer_reject_id": "Customer Reject Id (С)",
    "lead_reject_id": "Lead Reject Id (С)",
    "lead_source_id": "Lead Source Id (С)",
    # Кастомные поля
    "custom_ads_comp": "Назва кампанії (К)",
    "custom_id_srm": "ID з Facebook (К)",
    "custom_gorodstvaniya": "Громадянство (К)",
    "custom_age_": "Вік (К)",
    "custom_email": "Email (custom) (К)",
    "custom_yazik": "Мова (К)",
    "custom_urovenvladenwoo": "Рівень володіння (К)",
    "custom_schedule": "Розклад (К)",
    "custom_try_lessons": "Пробні уроки (К)",
}


def get_pipelines_and_statuses_from_api() -> Dict[int, Dict[str, Any]]:
    """
    Получить список статусов из AlfaCRM API.

    Returns:
        dict: {status_id: {"name": "...", "pipeline_id": ...}}
    """
    token = alfacrm_auth_get_token()
    base_url = os.getenv("ALFACRM_BASE_URL").rstrip('/')
    company_id = int(os.getenv("ALFACRM_COMPANY_ID"))

    # Получаем список статусов лидов
    try:
        url = f"{base_url}/v2api/lead-status/index"
        payload = {"branch_id": company_id}

        resp = requests.post(
            url,
            headers={"X-ALFACRM-TOKEN": token},
            json=payload,
            timeout=15
        )
        resp.raise_for_status()
        data = resp.json()

        statuses = data.get("items", [])

        # Создаем словарь статусов
        status_dict = {}
        for status in statuses:
            status_id = status.get("id")
            status_name = status.get("name", f"Статус {status_id}")
            pipeline_id = status.get("pipeline_id", 0)

            status_dict[status_id] = {
                "name": status_name,
                "pipeline_id": pipeline_id
            }

        return status_dict

    except Exception as e:
        print(f"  ⚠ Не вдалося отримати статуси через API: {e}")
        return {}


def build_pipelines_from_leads(leads: List[Dict[str, Any]], status_dict: Dict) -> Dict[int, Dict[str, Any]]:
    """
    Построить структуру воронок на основе реальных лидов с названиями статусов из API.

    Args:
        leads: Список лидов
        status_dict: Словарь статусов из API {status_id: {"name": "...", "pipeline_id": ...}}

    Returns:
        dict: {pipeline_id: {"name": "...", "statuses": {status_id: "name"}}}
    """
    print("\n[1/4] Побудова воронок на основі реальних лідів...")

    pipelines = defaultdict(lambda: {"name": "", "statuses": {}})
    pipeline_names = {}  # Имена воронок

    # Сначала получим имена воронок из API
    token = alfacrm_auth_get_token()
    base_url = os.getenv("ALFACRM_BASE_URL").rstrip('/')
    company_id = int(os.getenv("ALFACRM_COMPANY_ID"))

    try:
        url = f"{base_url}/v2api/pipeline/index"
        payload = {"branch_id": company_id}

        resp = requests.post(
            url,
            headers={"X-ALFACRM-TOKEN": token},
            json=payload,
            timeout=15
        )

        if resp.status_code == 200:
            data = resp.json()
            for pipeline in data.get("items", []):
                pipeline_id = pipeline.get("id")
                pipeline_name = pipeline.get("name", f"Воронка {pipeline_id}")
                pipeline_names[pipeline_id] = pipeline_name
    except:
        pass

    # Сначала определяем, в каких воронках есть ліди
    pipelines_with_leads = set()
    for lead in leads:
        pipeline_id = lead.get("pipeline_id")
        if pipeline_id is None:
            pipeline_id = 0
        pipelines_with_leads.add(pipeline_id)

    # Для каждой воронки с лидами создаем структуру со ВСЕМИ статусами из API
    for pipeline_id in pipelines_with_leads:
        pipeline_name = pipeline_names.get(pipeline_id, f"Воронка {pipeline_id}")

        pipelines[pipeline_id] = {
            "name": pipeline_name,
            "statuses": {}
        }

        # Добавляем специальный статус "Без статусу"
        pipelines[pipeline_id]["statuses"][-1] = "Без статусу"

        # Добавляем ВСЕ статусы из API
        for status_id, status_info in status_dict.items():
            status_name = status_info.get("name", f"Статус {status_id}")
            pipelines[pipeline_id]["statuses"][status_id] = status_name

    print(f"  ✓ Знайдено {len(pipelines)} воронок з лідами")
    for pid, pdata in pipelines.items():
        print(f"    - {pdata['name']}: {len(pdata['statuses'])} статусів")

    return dict(pipelines)


def get_all_leads() -> List[Dict[str, Any]]:
    """
    Получить ВСЕХ лидов из AlfaCRM включая АКТИВНЫХ и АРХИВНЫХ.

    Архивные ліди - это те у которых заполнено lead_reject_id.
    API не возвращает архивных лідів без фильтра, поэтому нужно:
    1. Получить всех активных (без lead_reject_id)
    2. Для каждой причины отказа получить архивных с пагинацией

    Returns:
        List[dict]: Список всех лидов
    """
    print("\n[2/4] Завантаження всіх лідів...")

    token = alfacrm_auth_get_token()
    base_url = os.getenv("ALFACRM_BASE_URL").rstrip('/')
    company_id = int(os.getenv("ALFACRM_COMPANY_ID"))

    all_leads_dict = {}  # Используем dict для хранения уникальных лидов по ID

    # ЭТАП 1: Получить всех АКТИВНЫХ лідів (без lead_reject_id)
    print("\n  [2.1] Завантаження активних лідів...")

    page = 1
    max_pages = 500
    pages_without_new = 0

    while page <= max_pages:
        url = f"{base_url}/v2api/customer/index"
        payload = {
            "branch_ids": [company_id],
            "page": page,
            "page_size": 500
        }

        try:
            resp = requests.post(url, headers={"X-ALFACRM-TOKEN": token}, json=payload, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            items = data.get("items", [])

            if not items:
                break

            new_leads = 0
            for lead in items:
                lead_id = lead.get("id")
                if lead_id not in all_leads_dict:
                    all_leads_dict[lead_id] = lead
                    new_leads += 1

            print(f"    Сторінка {page}: нових {new_leads}, всього {len(all_leads_dict)}")

            if new_leads == 0:
                pages_without_new += 1
                if pages_without_new >= 5:
                    break
            else:
                pages_without_new = 0

            page += 1
        except Exception as e:
            print(f"    ✗ Помилка: {e}")
            break

    active_count = len(all_leads_dict)
    print(f"  ✓ Активних лідів: {active_count}")

    # ЭТАП 2: Получить всех АРХИВНЫХ лідів (с lead_reject_id)
    print("\n  [2.2] Завантаження архівних лідів (за причинами відмов)...")

    # Получим список всех причин отказа
    url_rejects = f"{base_url}/v2api/lead-reject/index"
    try:
        resp = requests.post(url_rejects, headers={"X-ALFACRM-TOKEN": token}, json={}, timeout=15)
        lead_rejects = resp.json().get('items', [])
        print(f"    Знайдено {len(lead_rejects)} причин відмов")
        if lead_rejects:
            print(f"    Перші 3 причини: {[r.get('name') for r in lead_rejects[:3]]}")
    except Exception as e:
        print(f"    ✗ Не вдалося отримати причини відмов: {e}")
        lead_rejects = []

    # Для каждой причины отказа получаем лідів
    for idx, reject in enumerate(lead_rejects, 1):
        reject_id = reject.get('id')
        reject_name = reject.get('name', f'ID {reject_id}')

        page = 1
        pages_without_new = 0
        reject_leads = 0

        while page <= max_pages:
            payload = {
                "lead_reject_id": reject_id,
                # НЕ используем branch_ids - он конфликтует с lead_reject_id!
                "page": page,
                "page_size": 500
            }

            try:
                resp = requests.post(url, headers={"X-ALFACRM-TOKEN": token}, json=payload, timeout=15)
                items = resp.json().get("items", [])

                # Отладка для первой причины отказа на первой странице
                if idx == 1 and page == 1:
                    print(f"    [DEBUG] Первый запрос: reject_id={reject_id}, items={len(items)}")

                if not items:
                    break

                new_leads = 0
                for lead in items:
                    lead_id = lead.get("id")
                    if lead_id not in all_leads_dict:
                        all_leads_dict[lead_id] = lead
                        new_leads += 1
                        reject_leads += 1

                if new_leads == 0:
                    pages_without_new += 1
                    if pages_without_new >= 3:  # Для архива достаточно 3 пустых страниц
                        break
                else:
                    pages_without_new = 0

                page += 1
            except Exception as e:
                print(f"    ✗ Помилка для '{reject_name}': {e}")
                break

        if reject_leads > 0:
            print(f"    [{idx}/{len(lead_rejects)}] {reject_name[:40]}: +{reject_leads} лідів")

    archived_count = len(all_leads_dict) - active_count
    print(f"  ✓ Архівних лідів: {archived_count}")
    print(f"  ✓ ВСЬОГО лідів: {len(all_leads_dict)}")

    return list(all_leads_dict.values())


def format_value_for_excel(value: Any) -> str:
    """
    Форматировать значение для ячейки Excel.
    """
    if value is None or value == "":
        return ""

    if isinstance(value, list):
        if not value:
            return ""
        # Преобразуем список в строку через запятую
        return ", ".join(str(v) for v in value if v)

    if isinstance(value, bool):
        return "Так" if value else "Ні"

    if isinstance(value, dict):
        return str(value)

    return str(value)


def get_all_fields_from_leads(leads: List[Dict[str, Any]]) -> List[str]:
    """
    Получить список всех уникальных полей из лидов для создания заголовков таблицы.

    Args:
        leads: Список лидов

    Returns:
        List[str]: Список названий полей в нужном порядке
    """
    # Собираем все уникальные поля
    all_fields = set()
    for lead in leads:
        all_fields.update(lead.keys())

    # Определяем порядок важных полей
    important_fields = [
        "id", "name", "phone", "email",
        "created_at", "updated_at",
        "lead_status_id", "study_status_id", "pipeline_id",
        "balance", "paid_count", "paid_lesson_count",
    ]

    # Системные поля (не кастомные)
    system_fields = [f for f in all_fields if not f.startswith("custom_")]

    # Кастомные поля
    custom_fields = sorted([f for f in all_fields if f.startswith("custom_")])

    # Сначала важные поля, потом остальные системные, потом кастомные
    ordered_fields = []

    # Добавляем важные поля в нужном порядке
    for field in important_fields:
        if field in all_fields:
            ordered_fields.append(field)

    # Добавляем остальные системные поля (не важные)
    for field in sorted(system_fields):
        if field not in ordered_fields:
            ordered_fields.append(field)

    # Добавляем кастомные поля
    ordered_fields.extend(custom_fields)

    return ordered_fields


def export_to_excel_by_tabs(leads: List[Dict], pipelines: Dict, output_file: str):
    """
    Экспортировать лиды в Excel с вкладками по статусам.

    Args:
        leads: Список лидов
        pipelines: Информация о воронках и статусах
        output_file: Путь к выходному файлу
    """
    print(f"\n[3/4] Групування лідів по статусах...")

    # Группируем лиды по статусам
    leads_by_status = defaultdict(list)

    for lead in leads:
        lead_status = lead.get("lead_status_id")
        pipeline_id = lead.get("pipeline_id", 0)

        # Используем -1 для лидов без статуса (как в build_pipelines_from_leads)
        if lead_status is None:
            lead_status = -1

        # Ключ: (pipeline_id, status_id)
        key = (pipeline_id, lead_status)
        leads_by_status[key].append(lead)

    print(f"  ✓ Лідів згруповано по {len(leads_by_status)} статусах")

    # Получаем список всех полей
    all_fields = get_all_fields_from_leads(leads)
    print(f"  ✓ Знайдено {len(all_fields)} унікальних полів")

    # Создаем Excel файл
    print(f"\n[4/4] Створення Excel таблиці...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        sheet_count = 0
        first_sheet = True

        # Создаем вкладку для каждого статуса
        for pipeline_id, pipeline_data in sorted(pipelines.items()):
            pipeline_name = pipeline_data["name"]
            statuses = pipeline_data["statuses"]

            for status_id, status_name in sorted(statuses.items()):
                # Получаем лиды для этого статуса
                key = (pipeline_id, status_id)
                status_leads = leads_by_status.get(key, [])

                # Название вкладки
                if len(pipelines) > 1:
                    sheet_name = f"{status_name} - {pipeline_name}"
                else:
                    sheet_name = status_name

                # Excel имеет лимит на длину названия листа - 31 символ
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + "..."

                # Создаем вкладку
                if first_sheet:
                    # Используем стандартный лист для первой вкладки
                    ws = wb.active
                    ws.title = sheet_name
                    first_sheet = False
                else:
                    ws = wb.create_sheet(title=sheet_name)
                sheet_count += 1

                # Записываем заголовки
                for col_idx, field in enumerate(all_fields, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    # Используем украинский перевод или оригинальное название
                    cell.value = FIELD_TRANSLATIONS.get(field, field.replace("_", " ").title())
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Записываем данные лидов
                for row_idx, lead in enumerate(status_leads, start=2):
                    for col_idx, field in enumerate(all_fields, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = lead.get(field)
                        cell.value = format_value_for_excel(value)
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Настраиваем ширину столбцов
                for col_idx in range(1, len(all_fields) + 1):
                    col_letter = get_column_letter(col_idx)
                    # Устанавливаем ширину в зависимости от названия столбца
                    field = all_fields[col_idx - 1]
                    if field in ["id", "lead_status_id", "study_status_id", "pipeline_id"]:
                        ws.column_dimensions[col_letter].width = 10
                    elif field in ["name", "phone", "email"]:
                        ws.column_dimensions[col_letter].width = 25
                    elif field == "comment" or field == "note":
                        ws.column_dimensions[col_letter].width = 40
                    else:
                        ws.column_dimensions[col_letter].width = 18

                # Замораживаем первую строку
                ws.freeze_panes = "A2"

                print(f"  ✓ Створено вкладку '{sheet_name}': {len(status_leads)} лідів")

        # Сохраняем файл
        wb.save(output_file)
        print(f"\n  ✓ Файл збережено: {output_file}")
        print(f"  ✓ Створено {sheet_count} вкладок")

        # Статистика
        print(f"\n{'='*80}")
        print("СТАТИСТИКА ПО СТАТУСАХ:")
        print(f"{'='*80}\n")

        total_leads = 0
        for pipeline_id, pipeline_data in sorted(pipelines.items()):
            pipeline_name = pipeline_data["name"]
            statuses = pipeline_data["statuses"]

            for status_id, status_name in sorted(statuses.items()):
                key = (pipeline_id, status_id)
                count = len(leads_by_status.get(key, []))

                if count > 0:
                    total_leads += count
                    if len(pipelines) > 1:
                        full_name = f"{status_name} - {pipeline_name}"
                    else:
                        full_name = status_name

                    bar_length = min(int(count / 10), 50)
                    bar = "█" * bar_length
                    print(f"{full_name:50} │ {count:4} лідів │ {bar}")

        print(f"\n{'='*80}")
        print(f"ВСЬОГО ЛІДІВ: {total_leads}")
        print(f"{'='*80}\n")

    except ImportError:
        print("  ✗ Помилка: необхідна бібліотека openpyxl")
        print("  Встановіть: pip install openpyxl")
    except Exception as e:
        print(f"  ✗ Помилка при створенні Excel: {e}")
        import traceback
        traceback.print_exc()


def main():
    """
    Основна функція.
    """
    print("\n" + "="*80)
    print("ЕКСПОРТ ВСІХ ЛІДІВ З ALFACRM ПО ВКЛАДКАХ (СТАТУСАХ)")
    print("="*80)

    # 1. Получаем статусы из API
    status_dict = get_pipelines_and_statuses_from_api()
    print(f"  ✓ Отримано {len(status_dict)} статусів з API")

    # 2. Получаем всех лидов
    leads = get_all_leads()

    if not leads:
        print("\n❌ Не вдалося отримати лідів")
        return

    # 3. Строим воронки на основе реальных лидов
    pipelines = build_pipelines_from_leads(leads, status_dict)

    # 4. Экспортируем в Excel
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"Всі_ліди_по_вкладках_{timestamp}.xlsx"

    export_to_excel_by_tabs(leads, pipelines, output_file)

    print(f"✅ ГОТОВО! Експортовано {len(leads)} лідів")
    print(f"📄 Файл: {output_file}\n")


if __name__ == "__main__":
    main()
