import asyncio
import os
import uuid
import json
import logging
from datetime import datetime
from typing import Dict, Any, List
import openpyxl

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

from fastapi import FastAPI, BackgroundTasks, Request, Depends
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from dotenv import load_dotenv

from .progress import ProgressStore
from .connectors import meta as meta_conn
from .connectors import google_sheets as gs_conn
from .connectors import excel as excel_conn
from .mapping import load_mapping
from .connectors import crm as crm_tools
from .config_store import get_config_masked, set_config
from .connectors import crm as crm_conn
from .middleware.auth import verify_api_key
from .database import init_db, get_db
from .models import PipelineRun, RunLog
from .analytics_processor import AnalyticsProcessor
from .services import nethunt_tracking
from .services import alfacrm_tracking
from .services import meta_leads


load_dotenv()

app = FastAPI(title="Ads → Sheets → CRM")

# Initialize database on startup
@app.on_event("startup")
def startup_event():
    init_db()

# Security: Rate limiting
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Security: CORS protection
# TODO: Replace with your actual frontend domain in production
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:5173,http://localhost:8000,https://nasty-berries-stare.loca.lt,https://ecademy-api.loca.lt,https://*.railway.app").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "DELETE"],
    allow_headers=["Content-Type", "Authorization"],
)

# Security: Trusted host protection
# Railway domain pattern: facebook-ads-analytics-exporter-production.up.railway.app
ALLOWED_HOSTS = os.getenv("ALLOWED_HOSTS", "localhost,127.0.0.1,*.loca.lt,*.up.railway.app,*.railway.app").split(",")
# For Railway: disable strict host checking by allowing all hosts if RAILWAY_ENVIRONMENT is set
if os.getenv("RAILWAY_ENVIRONMENT"):
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=["*"])
else:
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=ALLOWED_HOSTS)

# Security: Add security headers middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    # Allow images from Facebook CDN for creative images
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https://*.fbcdn.net https://graph.facebook.com"
    )
    return response

progress = ProgressStore()

# Define paths for static files
WEB_DIST = os.path.join(os.path.dirname(os.path.dirname(__file__)), "web", "dist")
STATIC_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static")


async def run_analytics_task(job_id: str, params: Dict[str, Any]):
    """
    Background task для виконання аналітики через AnalyticsProcessor.
    """
    try:
        progress.update(job_id, 5, "Ініціалізація процесора")

        processor = AnalyticsProcessor(
            campaign_type=params["campaign_type"],
            date_start=params["date_start"],
            date_stop=params["date_stop"]
        )

        progress.update(job_id, 20, "Обробка кампаній...")
        logger.info(f"Starting analytics for {params['campaign_type']}")

        results = processor.process()

        progress.update(job_id, 90, "Збереження результатів")

        # Зберігаємо результати в Excel
        backend = os.getenv("STORAGE_BACKEND", "excel").lower()
        if backend == "excel":
            campaign_type = params["campaign_type"]
            if campaign_type == "students":
                excel_path = os.getenv("EXCEL_STUDENTS_PATH")
            else:
                excel_path = os.getenv("EXCEL_TEACHERS_PATH")

            if excel_path:
                from openpyxl import load_workbook, Workbook

                # Створюємо або завантажуємо Excel файл
                if os.path.exists(excel_path):
                    wb = load_workbook(excel_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    # Додаємо заголовки з першого результату
                    if results:
                        headers = list(results[0].keys())
                        ws.append(headers)

                # Додаємо дані
                for row_data in results:
                    row = [row_data.get(h) for h in results[0].keys()]
                    ws.append(row)

                wb.save(excel_path)
                progress.log(job_id, f"Збережено {len(results)} кампаній в {excel_path}")

        progress.update(job_id, 100, "done")
        logger.info(f"Analytics completed: {len(results)} campaigns processed")

    except Exception as e:
        logger.error(f"Analytics failed: {e}")
        progress.log(job_id, f"ERROR: {e}")
        progress.set_status(job_id, "error")


@app.post("/api/run")
@limiter.limit("5/minute")
async def run_analytics(
    request: Request,
    payload: Dict[str, Any],
    background_tasks: BackgroundTasks,
):
    """
    Новий endpoint для запуску аналітики на основі AnalyticsProcessor.

    Args:
        campaign_type: 'teachers' або 'students'
        date_start: Дата початку (YYYY-MM-DD)
        date_stop: Дата кінця (YYYY-MM-DD)
    """
    campaign_type = payload.get("campaign_type")
    date_start = payload.get("date_start")
    date_stop = payload.get("date_stop")

    # Валідація
    if campaign_type not in ["teachers", "students"]:
        return JSONResponse({"error": "campaign_type має бути 'teachers' або 'students'"}, status_code=400)

    if not date_start or not date_stop:
        return JSONResponse({"error": "Відсутні date_start або date_stop"}, status_code=400)

    try:
        _ = datetime.strptime(date_start, "%Y-%m-%d")
        _ = datetime.strptime(date_stop, "%Y-%m-%d")
    except ValueError:
        return JSONResponse({"error": "Невірний формат дати, використовуйте YYYY-MM-DD"}, status_code=400)

    job_id = str(uuid.uuid4())
    progress.init(job_id, title=f"Analytics: {campaign_type}")

    params = {
        "campaign_type": campaign_type,
        "date_start": date_start,
        "date_stop": date_stop,
    }
    background_tasks.add_task(run_analytics_task, job_id, params)
    return {"job_id": job_id}


@app.post("/api/start-job")
@limiter.limit("5/minute")
async def start_job(
    request: Request,
    payload: Dict[str, Any],
    background_tasks: BackgroundTasks,
):
    # payload expects: start_date, end_date (YYYY-MM-DD), sheet_id (optional)
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")
    sheet_id = payload.get("sheet_id") or os.getenv("GOOGLE_SHEET_ID")

    # Basic validation
    for key, val in {"start_date": start_date, "end_date": end_date}.items():
        if not val:
            return JSONResponse({"error": f"Відсутнє поле {key}"}, status_code=400)
    try:
        _ = datetime.strptime(start_date, "%Y-%m-%d")
        _ = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        return JSONResponse({"error": "Невірний формат дати, використовуйте YYYY-MM-DD"}, status_code=400)
    backend = os.getenv("STORAGE_BACKEND", "sheets").lower()
    if backend == "sheets" and not sheet_id:
        return JSONResponse({"error": "Відсутній Google Sheet ID"}, status_code=400)

    job_id = str(uuid.uuid4())
    progress.init(job_id, title="Pipeline run")

    params = {
        "start_date": start_date,
        "end_date": end_date,
        "sheet_id": sheet_id,
    }
    background_tasks.add_task(run_pipeline, job_id, params)
    return {"job_id": job_id}


@app.get("/api/events/{job_id}")
async def events(job_id: str):
    async def event_stream():
        last_index = 0
        while True:
            await asyncio.sleep(0.5)
            state = progress.get(job_id)
            if not state:
                # job unknown → end stream
                break
            logs = state["logs"]
            # Send incremental logs and current progress as SSE
            while last_index < len(logs):
                msg = logs[last_index]
                last_index += 1
                yield f"event: log\ndata: {msg}\n\n"
            yield (
                "event: progress\n"
                f"data: {{\"percent\": {state['percent']}, \"status\": \"{state['status']}\"}}\n\n"
            )
            if state["status"] in ("done", "error"):
                break

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.get("/api/inspect/excel-headers")
async def inspect_excel_headers():
    from .connectors import excel as ex
    # Just probe existing headers by reading first row; we won’t modify files here
    files = {
        "creatives": os.getenv("EXCEL_CREATIVES_PATH"),
        "students": os.getenv("EXCEL_STUDENTS_PATH"),
        "teachers": os.getenv("EXCEL_TEACHERS_PATH"),
    }
    result = {}
    for key, path in files.items():
        if not path or not os.path.exists(path):
            result[key] = {"error": "file_not_found", "path": path}
            continue
        try:
            # reuse openpyxl through helpers
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = {}
            for name in wb.sheetnames:
                ws = wb[name]
                row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
                headers = [str(c) if c is not None else '' for c in row1]
                sheets[name] = headers
            result[key] = {"path": path, "sheets": sheets}
        except Exception as e:
            result[key] = {"path": path, "error": str(e)}
    return result


@app.get("/api/inspect/nethunt/folders")
async def inspect_nethunt_folders():
    data = crm_tools.nethunt_list_folders()
    return data


@app.get("/api/inspect/nethunt/fields")
async def inspect_nethunt_fields(folder_id: str):
    data = crm_tools.nethunt_folder_fields(folder_id)
    return data


@app.get("/api/inspect/alfacrm/companies")
async def inspect_alfacrm_companies():
    data = crm_tools.alfacrm_list_companies()
    return data


@app.get("/health")
async def health_check():
    """Health check endpoint for Railway and monitoring"""
    return {"status": "healthy", "service": "ecademy-api"}


@app.get("/api/proxy-image")
async def proxy_image(url: str):
    """
    Proxy endpoint для загрузки изображений креативов.
    Обходит CORS ограничения Meta API.
    """
    import httpx

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, timeout=10.0)
            if response.status_code == 200:
                return StreamingResponse(
                    iter([response.content]),
                    media_type=response.headers.get("content-type", "image/jpeg")
                )
            else:
                return JSONResponse({"error": f"Failed to fetch image: {response.status_code}"}, status_code=response.status_code)
    except Exception as e:
        logger.error(f"Error proxying image: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/config")
@limiter.limit("10/minute")
async def get_config(request: Request):
    return get_config_masked()


@app.post("/api/config")
@limiter.limit("5/minute")
async def update_config(
    request: Request, payload: Dict[str, Any]
):
    # Never echo secrets back
    set_config(payload or {})
    return {"status": "ok"}


@app.get("/api/runs")
@limiter.limit("30/minute")
async def get_runs(
    request: Request,
    status: str = None,
    limit: int = 50,
    offset: int = 0
):
    """
    Get pipeline run history with optional filtering.

    Query params:
    - status: Filter by status (success, error, running)
    - limit: Maximum number of results (default: 50, max: 100)
    - offset: Pagination offset (default: 0)
    """
    try:
        # Limit maximum results
        limit = min(limit, 100)

        with get_db() as db:
            query = db.query(PipelineRun)

            # Apply status filter if provided
            if status:
                query = query.filter(PipelineRun.status == status)

            # Order by most recent first
            query = query.order_by(PipelineRun.start_time.desc())

            # Apply pagination
            runs = query.offset(offset).limit(limit).all()

            # Convert to dict
            result = []
            for run in runs:
                result.append({
                    "id": run.id,
                    "job_id": run.job_id,
                    "start_time": run.start_time.isoformat() if run.start_time else None,
                    "end_time": run.end_time.isoformat() if run.end_time else None,
                    "status": run.status,
                    "start_date": run.start_date,
                    "end_date": run.end_date,
                    "sheet_id": run.sheet_id,
                    "storage_backend": run.storage_backend,
                    "insights_count": run.insights_count,
                    "students_count": run.students_count,
                    "teachers_count": run.teachers_count,
                    "error_message": run.error_message
                })

            return {"runs": result, "count": len(result)}
    except Exception as e:
        return JSONResponse({"error": f"Помилка бази даних: {str(e)}"}, status_code=500)


@app.get("/api/runs/{run_id}")
@limiter.limit("30/minute")
async def get_run_details(request: Request, run_id: int):
    """Get detailed information about a specific run including logs."""
    try:
        with get_db() as db:
            run = db.query(PipelineRun).filter(PipelineRun.id == run_id).first()

            if not run:
                return JSONResponse({"error": "Запуск не знайдено"}, status_code=404)

            # Get logs for this run
            logs = db.query(RunLog).filter(RunLog.run_id == run_id).order_by(RunLog.timestamp).all()

            return {
                "run": {
                    "id": run.id,
                    "job_id": run.job_id,
                    "start_time": run.start_time.isoformat() if run.start_time else None,
                    "end_time": run.end_time.isoformat() if run.end_time else None,
                    "status": run.status,
                    "start_date": run.start_date,
                    "end_date": run.end_date,
                    "sheet_id": run.sheet_id,
                    "storage_backend": run.storage_backend,
                    "insights_count": run.insights_count,
                    "students_count": run.students_count,
                    "teachers_count": run.teachers_count,
                    "error_message": run.error_message
                },
                "logs": [
                    {
                        "id": log.id,
                        "timestamp": log.timestamp.isoformat() if log.timestamp else None,
                        "level": log.level,
                        "message": log.message
                    }
                    for log in logs
                ]
            }
    except Exception as e:
        return JSONResponse({"error": f"Помилка бази даних: {str(e)}"}, status_code=500)


@app.delete("/api/runs/{run_id}")
@limiter.limit("10/minute")
async def delete_run(request: Request, run_id: int):
    """Delete a pipeline run and its logs."""
    try:
        with get_db() as db:
            run = db.query(PipelineRun).filter(PipelineRun.id == run_id).first()

            if not run:
                return JSONResponse({"error": "Запуск не знайдено"}, status_code=404)

            # Delete associated logs first
            db.query(RunLog).filter(RunLog.run_id == run_id).delete()

            # Delete the run
            db.delete(run)
            db.commit()

            return {"success": True, "message": f"Запуск {run_id} видалено"}
    except Exception as e:
        logger.error(f"Error deleting run {run_id}: {e}")
        return JSONResponse({"error": f"Помилка видалення: {str(e)}"}, status_code=500)


async def enrich_students_with_alfacrm(students_from_excel: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Обогащает данные студентов из Excel свежими данными из AlfaCRM.

    Args:
        students_from_excel: Список студентов из Excel файла

    Returns:
        Обогащенный список студентов
    """
    if not (os.getenv("ALFACRM_BASE_URL") and os.getenv("ALFACRM_EMAIL")
            and os.getenv("ALFACRM_API_KEY") and os.getenv("ALFACRM_COMPANY_ID")):
        return students_from_excel

    try:
        alfacrm_students = []
        page = 1
        while True:
            data = crm_tools.alfacrm_list_students(page=page, page_size=200)
            items = data.get("items") or data.get("data") or data.get("list") or []
            if not isinstance(items, list):
                break

            for s in items:
                flat = {}
                if isinstance(s, dict):
                    for k, v in s.items():
                        if isinstance(v, (dict, list)):
                            try:
                                flat[k] = json.dumps(v, ensure_ascii=False)
                            except Exception:
                                flat[k] = str(v)
                        else:
                            flat[k] = v
                alfacrm_students.append(flat)

            if len(items) < 200:
                break
            page += 1

        alfacrm_lookup = {}
        for student in alfacrm_students:
            student_id = student.get("id") or student.get("student_id")
            email = student.get("email")
            if student_id:
                alfacrm_lookup[str(student_id)] = student
            if email:
                alfacrm_lookup[email] = student

        enriched = []
        for excel_student in students_from_excel:
            student_id = excel_student.get("id") or excel_student.get("student_id")
            email = excel_student.get("email") or excel_student.get("Email")

            crm_data = None
            if student_id and str(student_id) in alfacrm_lookup:
                crm_data = alfacrm_lookup[str(student_id)]
            elif email and email in alfacrm_lookup:
                crm_data = alfacrm_lookup[email]

            if crm_data:
                merged = {**excel_student, **crm_data}
            else:
                merged = excel_student

            enriched.append(merged)

        return enriched

    except Exception as e:
        print(f"Попередження: збагачення даних з AlfaCRM не вдалося: {e}")
        return students_from_excel


@app.post("/api/save-run-history")
@limiter.limit("30/minute")
async def save_run_history(request: Request, payload: Dict[str, Any]):
    """
    Зберігає історію запуску в базу даних.

    Args:
        start_date: YYYY-MM-DD
        end_date: YYYY-MM-DD
        insights_count: кількість insights
        students_count: кількість студентів
        teachers_count: кількість викладачів
        status: success/error
        error_message: опціонально
    """
    try:
        with get_db() as db:
            db_run = PipelineRun(
                job_id=str(uuid.uuid4()),
                start_time=datetime.utcnow(),
                end_time=datetime.utcnow(),
                start_date=payload.get("start_date"),
                end_date=payload.get("end_date"),
                storage_backend="meta_api",
                status=payload.get("status", "success"),
                insights_count=payload.get("insights_count", 0),
                students_count=payload.get("students_count", 0),
                teachers_count=payload.get("teachers_count", 0),
                error_message=payload.get("error_message")
            )
            db.add(db_run)
            db.commit()
            db.refresh(db_run)

            # Зберігаємо лог
            log_entry = RunLog(
                run_id=db_run.id,
                message=f"Meta API data fetched: {payload.get('insights_count', 0)} ads, {payload.get('students_count', 0)} students, {payload.get('teachers_count', 0)} teachers",
                level="info"
            )
            db.add(log_entry)
            db.commit()

            return {"success": True, "run_id": db_run.id}
    except Exception as e:
        logger.error(f"Error saving run history: {e}")
        return JSONResponse({"error": f"Помилка збереження історії: {str(e)}"}, status_code=500)


@app.get("/api/meta-data")
@limiter.limit("30/minute")
async def get_meta_data(request: Request, start_date: str = None, end_date: str = None):
    """
    Получить данные из Meta API для всех 3 вкладок за один запрос.

    Query params:
    - start_date: Начало периода (YYYY-MM-DD)
    - end_date: Конец периода (YYYY-MM-DD)

    Returns:
        {
            "ads": [...],      # Данные для вкладки РЕКЛАМА
            "students": [...], # Данные для вкладки СТУДЕНТИ
            "teachers": [...]  # Данные для вкладки ВЧИТЕЛІ
        }
    """
    try:
        meta_token = os.getenv("META_ACCESS_TOKEN")
        ad_account_id = os.getenv("META_AD_ACCOUNT_ID")

        if not meta_token or not ad_account_id:
            return JSONResponse({"error": "META credentials не налаштовані"}, status_code=400)

        if not start_date or not end_date:
            return JSONResponse({"error": "start_date та end_date обов'язкові"}, status_code=400)

        # 1) Получаем данные из Meta API (один раз для всех вкладок)
        logger.info(f"Fetching Meta data for period {start_date} - {end_date}")
        insights = meta_conn.fetch_insights(
            ad_account_id=ad_account_id,
            access_token=meta_token,
            date_from=start_date,
            date_to=end_date,
            level="ad"
        )

        logger.info(f"Received {len(insights)} insights from Meta API for period {start_date} - {end_date}")

        # 2) Получаем креативы (тексты и изображения)
        ad_ids = [insight.get("ad_id") for insight in insights if insight.get("ad_id")]
        creatives = meta_conn.fetch_ad_creatives(ad_ids, meta_token, ad_account_id) if ad_ids else {}

        # 3) Обогащаем insights креативами
        for insight in insights:
            ad_id = insight.get("ad_id")
            if ad_id and ad_id in creatives:
                creative_data = creatives[ad_id]
                insight["creative_title"] = creative_data.get("title", "")
                insight["creative_body"] = creative_data.get("body", "")
                insight["image_url"] = creative_data.get("image_url", "")
                insight["video_id"] = creative_data.get("video_id", "")

                # Debug logging для першого креативу
                if len(insights) > 0 and insight == insights[0]:
                    logger.info(f"DEBUG first creative: ad_id={ad_id}, image_url={creative_data.get('image_url')}, thumbnail={creative_data.get('thumbnail_url')}")

        # 4) Формируем данные для вкладки РЕКЛАМА
        ads_data = []
        for insight in insights:
            ads_data.append({
                "campaign_name": insight.get("campaign_name", ""),
                "campaign_id": insight.get("campaign_id", ""),
                "period": f"{insight.get('date_start', '')} - {insight.get('date_stop', '')}",
                "date_start": insight.get("date_start", ""),
                "date_stop": insight.get("date_stop", ""),
                "date_update": datetime.now().strftime("%Y-%m-%d"),
                "ad_name": insight.get("ad_name", ""),
                "creative_image": insight.get("image_url") or insight.get("thumbnail_url", ""),
                "creative_text": insight.get("creative_body", ""),
                "image_url": insight.get("image_url", ""),
                "thumbnail_url": insight.get("thumbnail_url", ""),
                "ctr": insight.get("ctr", 0),
                "cpl": "треба API",  # Cost per lead - требует leads API
                "cpm": insight.get("cpm", 0),
                "spend": insight.get("spend", 0),
                "leads_count": "треба API",
                "leads_target": "треба API",
                "leads_non_target": "треба API",
                "leads_no_answer": "треба API",
                "leads_in_progress": "треба API",
                "percent_target": "треба API",
                "percent_non_target": "треба API",
                "percent_no_answer": "треба API",
                "percent_in_progress": "треба API",
                "price_per_lead": "треба API",
                "price_per_target_lead": "треба API",
                "recommendation": ""
            })

        # 5) Формируем данные для вкладки СТУДЕНТИ з інтеграцією AlfaCRM tracking
        students_data = []

        # Отримуємо трекінг даних для студентів з AlfaCRM
        students_tracking = {}
        student_campaigns = {}
        student_index = {}

        try:
            meta_page_id = os.getenv("META_PAGE_ID")
            meta_page_token = os.getenv("META_PAGE_ACCESS_TOKEN")

            if meta_page_id and meta_page_token:
                # Фільтруємо лідів тільки для кампаній студентів
                keywords_students = os.getenv("CAMPAIGN_KEYWORDS_STUDENTS", "student,shkolnik").lower().split(",")

                all_campaigns = await meta_leads.get_leads_for_period(
                    page_id=meta_page_id,
                    page_token=meta_page_token,
                    start_date=start_date,
                    end_date=end_date
                )

                # Фільтруємо тільки кампанії студентів
                student_campaigns = {
                    cid: cdata for cid, cdata in all_campaigns.items()
                    if any(kw.strip() in cdata.get("campaign_name", "").lower() for kw in keywords_students)
                }

                # Завантажуємо всіх студентів з AlfaCRM для створення індексу
                from connectors.crm import alfacrm_list_students
                all_students = []
                page = 1
                while True:
                    try:
                        response = alfacrm_list_students(page=page, page_size=500)
                        students = response.get("items", [])
                        if not students:
                            break
                        all_students.extend(students)
                        total = response.get("total", 0)
                        if len(all_students) >= total:
                            break
                        page += 1
                    except Exception as e:
                        logger.error(f"Failed to load students page {page}: {e}")
                        break

                # Будуємо індекс студентів
                from services.alfacrm_tracking import build_student_index
                student_index = build_student_index(all_students)
                logger.info(f"Built student index with {len(student_index)} contact entries")

                # Трекінг через AlfaCRM з inference підходом
                students_tracking = await alfacrm_tracking.track_leads_by_campaigns(
                    campaigns_data=student_campaigns,
                    page_size=500
                )
                logger.info(f"Loaded student tracking for {len(students_tracking)} campaigns")
            else:
                logger.warning("META_PAGE_ID або META_PAGE_ACCESS_TOKEN не налаштовані - пропускаємо трекінг студентів")
        except Exception as e:
            logger.error(f"Failed to load student tracking: {e}")

        for insight in insights:
            campaign_name = insight.get("campaign_name", "").lower()
            campaign_id = insight.get("campaign_id", "")
            keywords_students = os.getenv("CAMPAIGN_KEYWORDS_STUDENTS", "student,shkolnik").lower().split(",")

            # Проверяем является ли кампания студенческой
            is_student_campaign = any(keyword.strip() in campaign_name for keyword in keywords_students)

            if is_student_campaign:
                # Отримуємо статистику воронки для цієї кампанії
                campaign_tracking = students_tracking.get(f"campaign_{campaign_id}", {})
                funnel_stats = campaign_tracking.get("funnel_stats", {})

                # Базові показники
                leads_count = funnel_stats.get("Кількість лідів", 0)
                budget = insight.get("spend", 0)

                # Маппінг статусів AlfaCRM на поля endpoint
                not_processed = funnel_stats.get("Не розібраний", 0)
                contact_established = funnel_stats.get("Вст контакт зацікавлений", 0)
                trial_scheduled = funnel_stats.get("Призначено пробне", 0)
                trial_completed = funnel_stats.get("Проведено пробне", 0)
                waiting_payment = funnel_stats.get("Чекаємо оплату", 0)
                purchased = funnel_stats.get("Отримана оплата", 0)

                # Недзвони (всі варіанти)
                no_answer = (
                    funnel_stats.get("Недодзвон", 0) +
                    funnel_stats.get("Недозвон 2", 0) +
                    funnel_stats.get("Недозвон 3", 0)
                )

                # Цільові/нецільові
                target_leads = contact_established  # Всі хто встановив контакт і зацікавлені
                non_target_leads = not_processed + no_answer  # Необроблені + недзвони

                # Розрахунок відсотків
                percent_target = round((target_leads / leads_count * 100), 2) if leads_count > 0 else 0
                percent_non_target = round((non_target_leads / leads_count * 100), 2) if leads_count > 0 else 0
                percent_contact = round((contact_established / leads_count * 100), 2) if leads_count > 0 else 0
                percent_conversion = round((purchased / leads_count * 100), 2) if leads_count > 0 else 0
                percent_no_answer = round((no_answer / leads_count * 100), 2) if leads_count > 0 else 0

                # Розрахунок для пробних уроків
                percent_trial_scheduled = round((trial_scheduled / leads_count * 100), 2) if leads_count > 0 else 0
                percent_trial_completed = round((trial_completed / leads_count * 100), 2) if leads_count > 0 else 0
                percent_trial_conversion = round((trial_completed / trial_scheduled * 100), 2) if trial_scheduled > 0 else 0
                conversion_trial_to_sale = round((purchased / trial_completed * 100), 2) if trial_completed > 0 else 0

                # Ціна за ліда
                price_per_lead = round((budget / leads_count), 2) if leads_count > 0 else 0
                price_per_target_lead = round((budget / target_leads), 2) if target_leads > 0 else 0

                students_data.append({
                    "campaign_name": insight.get("campaign_name", ""),
                    "campaign_link": f"https://facebook.com/ads/manager/campaigns/edit/{campaign_id}",
                    "analysis_date": datetime.now().strftime("%Y-%m-%d"),
                    "period": f"{start_date} - {end_date}",
                    "budget": budget,
                    "location": "треба API",  # Geo targeting требует дополнительный запрос
                    "leads_count": leads_count,
                    "leads_check": leads_count,  # Дублюємо для сумісності
                    "not_processed": not_processed,
                    "contact_established": contact_established,
                    "in_progress": funnel_stats.get("Розмовляли, чекаємо відповідь", 0),
                    "trial_scheduled": trial_scheduled,
                    "trial_completed": trial_completed,
                    "waiting_payment": waiting_payment,
                    "purchased": purchased,
                    "archive": funnel_stats.get("Зник після контакту", 0),
                    "no_answer": no_answer,
                    "archive_non_target": 0,  # TODO: визначити з клієнтом маппінг
                    "target_leads": target_leads,
                    "non_target_leads": non_target_leads,
                    "percent_target": percent_target,
                    "percent_non_target": percent_non_target,
                    "percent_contact": percent_contact,
                    "percent_in_progress": round((funnel_stats.get("Розмовляли, чекаємо відповідь", 0) / leads_count * 100), 2) if leads_count > 0 else 0,
                    "percent_conversion": percent_conversion,
                    "percent_archive": round((funnel_stats.get("Зник після контакту", 0) / leads_count * 100), 2) if leads_count > 0 else 0,
                    "percent_no_answer": percent_no_answer,
                    "price_per_lead": price_per_lead,
                    "price_per_target_lead": price_per_target_lead,
                    "notes": "",
                    "percent_trial_scheduled": percent_trial_scheduled,
                    "percent_trial_completed": percent_trial_completed,
                    "percent_trial_conversion": percent_trial_conversion,
                    "conversion_trial_to_sale": conversion_trial_to_sale
                })

        # 6) Формируем данные для вкладки ВЧИТЕЛІ з інтеграцією NetHunt tracking
        teachers_data = []

        # Отримуємо трекінг даних для вчителів з NetHunt
        teachers_tracking = {}
        teacher_campaigns = {}
        teacher_index = {}
        teacher_status_histories = {}

        try:
            meta_page_id = os.getenv("META_PAGE_ID")
            meta_page_token = os.getenv("META_PAGE_ACCESS_TOKEN")
            nh_folder = os.getenv("NETHUNT_FOLDER_ID")

            if meta_page_id and meta_page_token and nh_folder:
                # Фільтруємо лідів тільки для кампаній вчителів
                keywords_teachers = os.getenv("CAMPAIGN_KEYWORDS_TEACHERS", "teacher,vchitel").lower().split(",")

                all_campaigns = await meta_leads.get_leads_for_period(
                    page_id=meta_page_id,
                    page_token=meta_page_token,
                    start_date=start_date,
                    end_date=end_date
                )

                # Фільтруємо тільки кампанії вчителів
                teacher_campaigns = {
                    cid: cdata for cid, cdata in all_campaigns.items()
                    if any(kw.strip() in cdata.get("campaign_name", "").lower() for kw in keywords_teachers)
                }

                # Завантажуємо історію змін з NetHunt
                from services.nethunt_tracking import (
                    nethunt_get_record_changes,
                    extract_status_history,
                    build_teacher_index
                )
                from connectors.crm import nethunt_list_records

                # Завантажуємо всю історію змін
                all_changes = nethunt_get_record_changes(nh_folder)
                logger.info(f"Loaded {len(all_changes)} record changes from NetHunt")

                # Будуємо історії статусів
                record_ids = set(change.get("recordId") for change in all_changes if change.get("recordId"))
                for record_id in record_ids:
                    history = extract_status_history(record_id, all_changes)
                    if history:
                        teacher_status_histories[record_id] = history
                logger.info(f"Built status history for {len(teacher_status_histories)} teachers")

                # Завантажуємо поточні записи вчителів
                all_records = nethunt_list_records(nh_folder, limit=1000)
                teacher_index = build_teacher_index(all_records)
                logger.info(f"Built teacher index with {len(teacher_index)} contact entries")

                # Трекінг через NetHunt з real history підходом
                teachers_tracking = await nethunt_tracking.track_leads_by_campaigns(
                    campaigns_data=teacher_campaigns,
                    folder_id=nh_folder
                )
                logger.info(f"Loaded teacher tracking for {len(teachers_tracking)} campaigns")
            else:
                logger.warning("META_PAGE_ID, META_PAGE_ACCESS_TOKEN або NETHUNT_FOLDER_ID не налаштовані")
        except Exception as e:
            logger.error(f"Failed to load teacher tracking: {e}")

        for insight in insights:
            campaign_name = insight.get("campaign_name", "").lower()
            campaign_id = insight.get("campaign_id", "")
            keywords_teachers = os.getenv("CAMPAIGN_KEYWORDS_TEACHERS", "teacher,vchitel").lower().split(",")

            # Проверяем является ли кампания для викладачів
            is_teacher_campaign = any(keyword.strip() in campaign_name for keyword in keywords_teachers)

            if is_teacher_campaign:
                # Отримуємо статистику воронки для цієї кампанії
                campaign_tracking = teachers_tracking.get(f"campaign_{campaign_id}", {})
                funnel_stats = campaign_tracking.get("funnel_stats", {})

                # Базові показники
                leads_count = funnel_stats.get("Кількість лідів", 0)
                budget = insight.get("spend", 0)

                # Маппінг статусів NetHunt на поля endpoint (основні 9 статусів)
                new_leads = funnel_stats.get("Нові", 0)
                contact_established = funnel_stats.get("Контакт встановлено", 0)
                qualified = funnel_stats.get("Кваліфіковані", 0)
                interview_scheduled = funnel_stats.get("Співбесіда призначена", 0)
                interview_completed = funnel_stats.get("Співбесіда проведена", 0)
                offer_sent = funnel_stats.get("Офер відправлено", 0)
                hired = funnel_stats.get("Найнято", 0)
                rejected = funnel_stats.get("Відмова", 0)
                no_answer = funnel_stats.get("Недзвін", 0)

                # Цільові/нецільові (базова логіка - можна уточнити у клієнта)
                target_leads = contact_established  # Всі хто встановив контакт
                non_target_leads = new_leads + no_answer  # Нові + недзвін

                # Конверсії
                conv_lead_to_interview = round((interview_completed / leads_count * 100), 2) if leads_count > 0 else 0
                conv_interview_to_hire = round((hired / interview_completed * 100), 2) if interview_completed > 0 else 0

                # Відсотки
                percent_target = round((target_leads / leads_count * 100), 2) if leads_count > 0 else 0
                percent_non_target = round((non_target_leads / leads_count * 100), 2) if leads_count > 0 else 0

                # Ціна за ліда
                price_per_lead = round((budget / leads_count), 2) if leads_count > 0 else 0
                price_per_target_lead = round((budget / target_leads), 2) if target_leads > 0 else 0

                teachers_data.append({
                    "campaign_name": insight.get("campaign_name", ""),
                    "campaign_link": f"https://facebook.com/ads/manager/campaigns/edit/{campaign_id}",
                    "analysis_date": datetime.now().strftime("%Y-%m-%d"),
                    "period": f"{start_date} - {end_date}",
                    "budget": budget,
                    "location": "треба API",  # Geo targeting потребує окремий запрос
                    "leads_count": leads_count,
                    "leads_check": leads_count,
                    "new_leads": new_leads,
                    "contact_established": contact_established,
                    "qualified": qualified,
                    "interview_scheduled": interview_scheduled,
                    "interview_completed": interview_completed,
                    "offer_sent": offer_sent,
                    "hired": hired,
                    "rejected": rejected,
                    "no_answer": no_answer,
                    "target_leads": target_leads,
                    "non_target_leads": non_target_leads,
                    "conversion_hired": round((hired / leads_count * 100), 2) if leads_count > 0 else 0,
                    "conversion_qualified": round((qualified / leads_count * 100), 2) if leads_count > 0 else 0,
                    "conversion_lead_to_interview": conv_lead_to_interview,
                    "conversion_interview_to_hire": conv_interview_to_hire,
                    "percent_target": percent_target,
                    "percent_non_target": percent_non_target,
                    "price_per_lead": price_per_lead,
                    "price_per_target_lead": price_per_target_lead,
                    "campaign_status": "active"  # Можна додати логіку визначення статусу
                })

        # Додаємо metadata для кольорової маркіровки стовпців в UI
        column_metadata = _get_column_metadata()

        # Витягуємо телефони лідів з інформацією про passed/current статуси
        lead_phones_students = {}
        lead_phones_teachers = {}

        try:
            # Студенти: витягуємо телефони з AlfaCRM inference підходом
            if student_campaigns and student_index:
                lead_phones_students = _extract_lead_phones_with_status_students(
                    campaigns_data=student_campaigns,
                    student_index=student_index,
                    analysis_date=datetime.now().strftime("%Y-%m-%d")
                )
                logger.info(f"Extracted phone data for {len(lead_phones_students)} student campaigns")
        except Exception as e:
            logger.error(f"Failed to extract student phone data: {e}")

        try:
            # Вчителі: витягуємо телефони з NetHunt real history
            if teacher_campaigns and teacher_index and teacher_status_histories:
                lead_phones_teachers = _extract_lead_phones_with_status_teachers(
                    campaigns_data=teacher_campaigns,
                    teacher_index=teacher_index,
                    status_histories=teacher_status_histories,
                    analysis_date=datetime.now().strftime("%Y-%m-%d")
                )
                logger.info(f"Extracted phone data for {len(lead_phones_teachers)} teacher campaigns")
        except Exception as e:
            logger.error(f"Failed to extract teacher phone data: {e}")

        return {
            "ads": ads_data,
            "students": students_data,
            "teachers": teachers_data,
            "column_metadata": column_metadata,
            "lead_phones": {
                "students": lead_phones_students,
                "teachers": lead_phones_teachers
            },
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "period": f"{start_date} - {end_date}"
        }

    except Exception as e:
        logger.error(f"Error fetching Meta data: {e}")
        return JSONResponse({"error": f"Помилка отримання даних: {str(e)}"}, status_code=500)


def _extract_lead_phones_with_status_students(
    campaigns_data: Dict[str, Dict[str, Any]],
    student_index: Dict[str, Dict[str, Any]],
    analysis_date: str
) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    """
    Витягує телефони студентів з інформацією про їх статус (passed/current).

    Args:
        campaigns_data: Дані кампаній з Meta leads API
        student_index: Індекс студентів для матчингу
        analysis_date: Дата аналізу для визначення passed/current

    Returns:
        {
            "campaign_123": {
                "Не розібраний": [
                    {"phone": "+380...", "status": "passed"},
                    {"phone": "+380...", "status": "current"}
                ],
                ...
            }
        }
    """
    from .services.alfacrm_tracking import extract_lead_contacts, normalize_contact
    from .services.lead_journey_recovery import recover_lead_journey, ALFACRM_STATUS_NAMES

    result = {}

    for campaign_id, campaign_data in campaigns_data.items():
        campaign_leads = campaign_data.get("leads", [])
        if not campaign_leads:
            continue

        # Словник для зберігання лідів по статусах
        status_phones = {}

        for lead in campaign_leads:
            phone, email = extract_lead_contacts(lead)
            if not phone and not email:
                continue

            # Знаходимо студента в CRM
            student = None
            if phone and phone in student_index:
                student = student_index[phone]
            elif email and email in student_index:
                student = student_index[email]

            if not student:
                # Лід не знайдено в CRM - вважаємо "Не розібраний" і current
                status_name = "Не розібраний"
                if status_name not in status_phones:
                    status_phones[status_name] = []
                status_phones[status_name].append({
                    "phone": phone or email,
                    "status": "current"
                })
                continue

            # Визначаємо поточний статус студента
            current_status_id = student.get("lead_status_id")
            if not current_status_id:
                continue

            # Отримуємо journey - всі статуси через які пройшов лід
            journey = recover_lead_journey(current_status_id)

            # Останній статус в journey - поточний
            current_status_name = ALFACRM_STATUS_NAMES.get(current_status_id)

            # Додаємо телефон в КОЖЕН статус через який пройшов лід
            for status_id in journey:
                status_name = ALFACRM_STATUS_NAMES.get(status_id)
                if not status_name:
                    continue

                if status_name not in status_phones:
                    status_phones[status_name] = []

                # Визначаємо passed або current
                is_current = (status_id == current_status_id)

                status_phones[status_name].append({
                    "phone": phone or email,
                    "status": "current" if is_current else "passed"
                })

        result[campaign_id] = status_phones

    return result


def _extract_lead_phones_with_status_teachers(
    campaigns_data: Dict[str, Dict[str, Any]],
    teacher_index: Dict[str, Dict[str, Any]],
    status_histories: Dict[str, List[Dict[str, Any]]],
    analysis_date: str
) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    """
    Витягує телефони вчителів з інформацією про їх статус (passed/current).

    Використовує РЕАЛЬНУ історію статусів з NetHunt API.

    Args:
        campaigns_data: Дані кампаній з Meta leads API
        teacher_index: Індекс вчителів для матчингу
        status_histories: Історії змін статусів з NetHunt
        analysis_date: Дата аналізу для визначення passed/current

    Returns:
        {
            "campaign_123": {
                "Нові": [
                    {"phone": "+380...", "status": "passed"},
                    {"phone": "+380...", "status": "current"}
                ],
                ...
            }
        }
    """
    from .services.nethunt_tracking import extract_lead_contacts, map_nethunt_status_to_column

    result = {}

    for campaign_id, campaign_data in campaigns_data.items():
        campaign_leads = campaign_data.get("leads", [])
        if not campaign_leads:
            continue

        # Словник для зберігання лідів по статусах
        status_phones = {}

        for lead in campaign_leads:
            phone, email = extract_lead_contacts(lead)
            if not phone and not email:
                continue

            # Знаходимо вчителя в CRM
            teacher_record = None
            if phone and phone in teacher_index:
                teacher_record = teacher_index[phone]
            elif email and email in teacher_index:
                teacher_record = teacher_index[email]

            if not teacher_record:
                # Лід не знайдено в CRM - вважаємо "Нові" і current
                status_name = "Нові"
                if status_name not in status_phones:
                    status_phones[status_name] = []
                status_phones[status_name].append({
                    "phone": phone or email,
                    "status": "current"
                })
                continue

            # Отримуємо реальну історію статусів з NetHunt
            record_id = teacher_record.get("id")
            history = status_histories.get(record_id, [])

            if not history:
                # Якщо немає історії - використовуємо поточний статус
                from .services.nethunt_tracking import extract_status_from_record
                status_name = extract_status_from_record(teacher_record)
                column_name = map_nethunt_status_to_column(status_name)

                if column_name not in status_phones:
                    status_phones[column_name] = []
                status_phones[column_name].append({
                    "phone": phone or email,
                    "status": "current"
                })
                continue

            # Визначаємо останній статус (поточний)
            last_change = history[-1] if history else None
            current_status = last_change.get("new_status", "").lower() if last_change else None
            current_column = map_nethunt_status_to_column(current_status)

            # Збираємо всі унікальні статуси через які пройшов лід
            unique_statuses = set()
            for change in history:
                new_status = change.get("new_status", "").lower() if change.get("new_status") else None
                if new_status:
                    column_name = map_nethunt_status_to_column(new_status)
                    unique_statuses.add(column_name)

            # Додаємо телефон в КОЖЕН статус через який пройшов лід
            for status_col in unique_statuses:
                if status_col not in status_phones:
                    status_phones[status_col] = []

                # Визначаємо passed або current
                is_current = (status_col == current_column)

                status_phones[status_col].append({
                    "phone": phone or email,
                    "status": "current" if is_current else "passed"
                })

        result[campaign_id] = status_phones

    return result


def _get_column_metadata() -> Dict[str, Dict[str, str]]:
    """
    Генерує metadata для кольорової маркіровки стовпців в UI.

    Returns:
        {
            "students": {"field_name": "meta|crm|formula", ...},
            "teachers": {"field_name": "meta|crm|formula", ...},
            "ads": {"field_name": "meta|crm|formula", ...}
        }
    """
    # Загальні поля Meta для всіх вкладок
    meta_fields = {
        "campaign_name", "campaign_link", "campaign_id", "period", "budget",
        "analysis_date", "location", "date_start", "date_stop", "date_update",
        "ad_name", "creative_image", "creative_text", "image_url", "thumbnail_url",
        "ctr", "cpm", "spend", "impressions", "clicks", "cpc"
    }

    # Студенти
    students_crm = {
        "not_processed", "contact_established", "in_progress",
        "trial_scheduled", "trial_completed", "waiting_payment",
        "purchased", "archive", "no_answer", "archive_non_target",
        "leads_count", "leads_check"
    }
    students_formula = {
        "target_leads", "non_target_leads", "percent_target", "percent_non_target",
        "percent_contact", "percent_in_progress", "percent_conversion",
        "percent_archive", "percent_no_answer", "price_per_lead",
        "price_per_target_lead", "notes", "percent_trial_scheduled",
        "percent_trial_completed", "percent_trial_conversion",
        "conversion_trial_to_sale"
    }

    # Вчителі
    teachers_crm = {
        "new_leads", "contact_established", "qualified",
        "interview_scheduled", "interview_completed", "offer_sent",
        "hired", "rejected", "no_answer", "leads_count", "leads_check"
    }
    teachers_formula = {
        "target_leads", "non_target_leads", "conversion_hired",
        "conversion_qualified", "conversion_lead_to_interview",
        "conversion_interview_to_hire", "percent_target", "percent_non_target",
        "price_per_lead", "price_per_target_lead", "campaign_status"
    }

    # Реклама
    ads_formula = {
        "cpl", "leads_count", "leads_target", "leads_non_target",
        "leads_no_answer", "leads_in_progress", "percent_target",
        "percent_non_target", "percent_no_answer", "percent_in_progress",
        "price_per_lead", "price_per_target_lead", "recommendation"
    }

    # Будуємо словники metadata для кожної вкладки
    students_metadata = {}
    for field in meta_fields:
        students_metadata[field] = "meta"
    for field in students_crm:
        students_metadata[field] = "crm"
    for field in students_formula:
        students_metadata[field] = "formula"

    teachers_metadata = {}
    for field in meta_fields:
        teachers_metadata[field] = "meta"
    for field in teachers_crm:
        teachers_metadata[field] = "crm"
    for field in teachers_formula:
        teachers_metadata[field] = "formula"

    ads_metadata = {}
    for field in meta_fields:
        ads_metadata[field] = "meta"
    for field in ads_formula:
        ads_metadata[field] = "formula"

    return {
        "students": students_metadata,
        "teachers": teachers_metadata,
        "ads": ads_metadata
    }


def _apply_column_color_coding(ws, headers, data_type="students"):
    """
    Застосовує цветову маркіровку стовпців на основі їх типу.

    Голубий (AddCDE) = Дані з Meta Ads
    Розовий (FFB6C1) = Дані з CRM
    Зелений (90EE90) = Формули та розрахунки

    Args:
        ws: Worksheet openpyxl
        headers: Список заголовків стовпців
        data_type: "students" | "teachers" | "ads"
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    # Визначаємо кольори
    COLOR_META = "AddCDE"      # Голубий - дані з Meta
    COLOR_CRM = "FFB6C1"       # Розовий - дані з CRM
    COLOR_FORMULA = "90EE90"   # Зелений - формули та розрахунки

    # Визначаємо які поля належать до якої категорії
    meta_fields = {
        "campaign_name", "campaign_link", "campaign_id", "period", "budget",
        "analysis_date", "location", "date_start", "date_stop", "date_update",
        "ad_name", "creative_image", "creative_text", "image_url", "thumbnail_url",
        "ctr", "cpm", "spend", "impressions", "clicks", "cpc"
    }

    if data_type == "students":
        crm_fields = {
            "not_processed", "contact_established", "in_progress",
            "trial_scheduled", "trial_completed", "waiting_payment",
            "purchased", "archive", "no_answer", "archive_non_target",
            "leads_count", "leads_check"
        }
        formula_fields = {
            "target_leads", "non_target_leads", "percent_target", "percent_non_target",
            "percent_contact", "percent_in_progress", "percent_conversion",
            "percent_archive", "percent_no_answer", "price_per_lead",
            "price_per_target_lead", "notes", "percent_trial_scheduled",
            "percent_trial_completed", "percent_trial_conversion",
            "conversion_trial_to_sale"
        }
    elif data_type == "teachers":
        crm_fields = {
            "new_leads", "contact_established", "qualified",
            "interview_scheduled", "interview_completed", "offer_sent",
            "hired", "rejected", "no_answer", "leads_count", "leads_check"
        }
        formula_fields = {
            "target_leads", "non_target_leads", "conversion_hired",
            "conversion_qualified", "conversion_lead_to_interview",
            "conversion_interview_to_hire", "percent_target", "percent_non_target",
            "price_per_lead", "price_per_target_lead", "campaign_status"
        }
    else:  # ads
        crm_fields = set()
        formula_fields = {
            "cpl", "leads_count", "leads_target", "leads_non_target",
            "leads_no_answer", "leads_in_progress", "percent_target",
            "percent_non_target", "percent_no_answer", "percent_in_progress",
            "price_per_lead", "price_per_target_lead", "recommendation"
        }

    # Застосовуємо кольори до заголовків
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)

        # Визначаємо колір на основі типу поля
        if header in meta_fields:
            fill_color = COLOR_META
        elif header in crm_fields:
            fill_color = COLOR_CRM
        elif header in formula_fields:
            fill_color = COLOR_FORMULA
        else:
            fill_color = "FFFFFF"  # Білий за замовчуванням

        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


@app.post("/api/export-meta-excel")
@limiter.limit("10/minute")
async def export_meta_excel(request: Request, payload: Dict[str, Any]):
    """
    Експорт даних з всіх 3 вкладок (Реклама, Студенти, Вчителі) в один Excel файл.
    З цветовою маркіровкою стовпців:
    - Голубий = Дані з Meta Ads
    - Розовий = Дані з CRM
    - Зелений = Формули та розрахунки

    Body:
        {
            "ads": [...],
            "students": [...],
            "teachers": [...]
        }
    """
    from fastapi.responses import FileResponse
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    try:
        ads_data = payload.get("ads", [])
        students_data = payload.get("students", [])
        teachers_data = payload.get("teachers", [])

        wb = Workbook()

        # Лист 1: Реклама
        ws_ads = wb.active
        ws_ads.title = "Реклама"

        if ads_data:
            # Заголовки для реклами
            ads_headers = list(ads_data[0].keys())
            ws_ads.append(ads_headers)

            # Застосовуємо цветову маркіровку
            _apply_column_color_coding(ws_ads, ads_headers, data_type="ads")

            # Дані
            for row_data in ads_data:
                ws_ads.append(list(row_data.values()))

        # Лист 2: Студенти
        ws_students = wb.create_sheet("Студенти")
        if students_data:
            students_headers = list(students_data[0].keys())
            ws_students.append(students_headers)

            for col_idx, header in enumerate(students_headers, 1):
                cell = ws_students.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_data in students_data:
                ws_students.append(list(row_data.values()))

        # Лист 3: Вчителі
        ws_teachers = wb.create_sheet("Вчителі")
        if teachers_data:
            teachers_headers = list(teachers_data[0].keys())
            ws_teachers.append(teachers_headers)

            for col_idx, header in enumerate(teachers_headers, 1):
                cell = ws_teachers.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_data in teachers_data:
                ws_teachers.append(list(row_data.values()))

        # Зберігаємо у тимчасовий файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f"_meta_data_{timestamp}.xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        logger.info(f"Excel file created: {temp_file.name}")

        return FileResponse(
            path=temp_file.name,
            filename=f"ecademy_meta_data_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        return JSONResponse({"error": f"Помилка експорту: {str(e)}"}, status_code=500)


@app.get("/api/students")
@limiter.limit("30/minute")
async def get_students(request: Request, start_date: str = None, end_date: str = None, enrich: bool = True):
    """
    Get students data with all 33 fields from Students Analysis structure.

    Query params:
    - start_date: Filter by analysis date (YYYY-MM-DD)
    - end_date: Filter by analysis date (YYYY-MM-DD)
    - enrich: Enrich with fresh AlfaCRM data (default: True)
    """
    try:
        backend = os.getenv("STORAGE_BACKEND", "excel").lower()

        if backend == "excel":
            excel_path = os.getenv("EXCEL_STUDENTS_PATH")
            if not excel_path or not os.path.exists(excel_path):
                return JSONResponse({"error": "Файл студентів не знайдено"}, status_code=404)

            # Read from Excel
            from openpyxl import load_workbook
            mapping = load_mapping()
            sheet_name = mapping.get("students", {}).get("sheet_name", "Students")

            wb = load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                return JSONResponse({"error": f"Аркуш '{sheet_name}' не знайдено"}, status_code=404)

            ws = wb[sheet_name]

            # Get headers from first row
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not headers_row:
                return JSONResponse({"error": "Заголовки не знайдено"}, status_code=404)

            headers = [str(h) if h is not None else '' for h in headers_row]

            # Read all data rows
            students = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue

                student_dict = {}
                for idx, header in enumerate(headers):
                    value = row[idx] if idx < len(row) else None
                    # Convert datetime to string
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    student_dict[header] = value

                students.append(student_dict)

            wb.close()

            # Filter by date if provided
            if start_date or end_date:
                filtered = []
                for s in students:
                    date_str = s.get("Дата аналізу", "")
                    if date_str:
                        # Simple date comparison
                        if start_date and str(date_str) < start_date:
                            continue
                        if end_date and str(date_str) > end_date:
                            continue
                    filtered.append(s)
                students = filtered

            # Enrich with AlfaCRM data if requested
            if enrich:
                students = await enrich_students_with_alfacrm(students)

            return {"students": students, "count": len(students)}
        else:
            # Google Sheets backend - to be implemented
            return JSONResponse({"error": "Google Sheets backend поки не підтримується"}, status_code=501)

    except Exception as e:
        return JSONResponse({"error": f"Помилка читання даних: {str(e)}"}, status_code=500)


def _add_students_charts(ws):
    """Добавляет графики для данных студентов."""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    if ws.max_row < 2:
        return

    headers = [cell.value for cell in ws[1]]

    # График 1: Бар-чарт для цільових/нецільових лідів
    try:
        target_col_idx = headers.index("Цільові ліди") + 1 if "Цільові ліди" in headers else None
        non_target_col_idx = headers.index("Не цільові ліди") + 1 if "Не цільові ліди" in headers else None

        if target_col_idx and non_target_col_idx:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Розподіл цільових та нецільових лідів"
            chart.y_axis.title = "Кількість лідів"
            chart.x_axis.title = "Креатив"

            data = Reference(ws, min_col=target_col_idx, min_row=1, max_row=ws.max_row, max_col=non_target_col_idx)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25
            ws.add_chart(chart, f"A{ws.max_row + 3}")
    except Exception as e:
        logger.warning(f"Не вдалося створити графік цільових лідів: {e}")

    # График 2: Pie chart для конверсії
    try:
        conversion_col_idx = headers.index("% конверсія") + 1 if "% конверсія" in headers else None

        if conversion_col_idx and ws.max_row <= 10:
            pie = PieChart()
            pie.title = "Конверсія по креативах"
            labels = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 10))
            data = Reference(ws, min_col=conversion_col_idx, min_row=2, max_row=min(ws.max_row, 10))
            pie.add_data(data)
            pie.set_categories(labels)
            pie.height = 12
            pie.width = 18
            ws.add_chart(pie, f"N{ws.max_row + 3}")
    except Exception as e:
        logger.warning(f"Не вдалося створити pie chart конверсії: {e}")


def _add_ads_charts(ws):
    """Добавляет графики для рекламных данных."""
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    if ws.max_row < 2:
        return

    headers = [cell.value for cell in ws[1]]

    # График 1: Line chart для spend, impressions, clicks
    try:
        spend_col_idx = headers.index("spend") + 1 if "spend" in headers else None
        impressions_col_idx = headers.index("impressions") + 1 if "impressions" in headers else None
        clicks_col_idx = headers.index("clicks") + 1 if "clicks" in headers else None

        if spend_col_idx:
            chart = LineChart()
            chart.title = "Витрати та показники по датах"
            chart.style = 13
            chart.y_axis.title = "Значення"
            chart.x_axis.title = "Дата"

            data = Reference(ws, min_col=spend_col_idx, min_row=1, max_row=min(ws.max_row, 30))
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 30))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25
            ws.add_chart(chart, f"A{ws.max_row + 3}")
    except Exception as e:
        logger.warning(f"Не вдалося створити line chart: {e}")

    # График 2: Bar chart для CTR и CPC
    try:
        ctr_col_idx = headers.index("ctr") + 1 if "ctr" in headers else None
        cpc_col_idx = headers.index("cpc") + 1 if "cpc" in headers else None

        if ctr_col_idx and cpc_col_idx:
            chart = BarChart()
            chart.type = "col"
            chart.title = "CTR та CPC по оголошеннях"
            chart.y_axis.title = "Значення"
            chart.x_axis.title = "Оголошення"

            data = Reference(ws, min_col=ctr_col_idx, min_row=1, max_row=min(ws.max_row, 20), max_col=cpc_col_idx)
            cats = Reference(ws, min_col=8, min_row=2, max_row=min(ws.max_row, 20))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25
            ws.add_chart(chart, f"N{ws.max_row + 3}")
    except Exception as e:
        logger.warning(f"Не вдалося створити bar chart для CTR/CPC: {e}")


@app.post("/api/download-excel")
@limiter.limit("5/minute")
async def download_excel(request: Request, payload: Dict[str, Any]):
    """
    Експорт даних у Excel з підтримкою різних типів даних.

    Args:
        data_type: "ads" | "students" | "teachers"
        start_date: YYYY-MM-DD
        end_date: YYYY-MM-DD
    """
    from fastapi.responses import FileResponse
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    data_type = payload.get("data_type", "ads")
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")

    # Валідація data_type
    if data_type not in ["ads", "students", "teachers"]:
        return JSONResponse(
            {"error": f"Невірний тип даних. Дозволені: 'ads', 'students', 'teachers'. Отримано: '{data_type}'"},
            status_code=400
        )

    try:
        wb = Workbook()
        ws = wb.active

        if data_type == "students":
            ws.title = "Students"

            # Отримуємо дані студентів через існуючий endpoint
            excel_path = os.getenv("EXCEL_STUDENTS_PATH")
            if not excel_path or not os.path.exists(excel_path):
                return JSONResponse({"error": "Файл студентів не знайдено"}, status_code=404)

            # Читаємо з Excel
            mapping = load_mapping()
            sheet_name = mapping.get("students", {}).get("sheet_name", "Students")

            source_wb = openpyxl.load_workbook(excel_path, data_only=True)
            if sheet_name not in source_wb.sheetnames:
                return JSONResponse({"error": f"Аркуш '{sheet_name}' не знайдено"}, status_code=404)

            source_ws = source_wb[sheet_name]

            # Копіюємо всі дані з форматуванням
            for row_idx, row in enumerate(source_ws.iter_rows(values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    target_cell = ws.cell(row=row_idx, column=col_idx)
                    target_cell.value = cell.value

                    # Форматування заголовків
                    if row_idx == 1:
                        target_cell.font = Font(bold=True, size=11)
                        target_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        target_cell.font = Font(bold=True, color="FFFFFF", size=11)
                        target_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Автоширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Применяем числовые форматы для процентов и валюты
            percent_columns = ["% цільових лідів", "% не цільових лідів", "% Встан. контакт",
                             "% В опрацюванні (ЦА)", "% конверсія", "% архів", "% недозвон",
                             "% Назначений пробний", "%\nПроведений пробний від загальних лідів\n(ЦА)",
                             "%\nПроведений пробний від назначених пробних", "Конверсія з проведеного пробного в продаж"]

            currency_columns = ["Витрачений бюджет в $", "Ціна / ліда", "Ціна / цільового ліда"]

            # Находим индексы колонок
            headers = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(headers, 1):
                if header in percent_columns:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = '0.00%'
                elif header in currency_columns:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = '$#,##0.00'

            source_wb.close()

            # Добавляем графики для студентов
            _add_students_charts(ws)

            filename = f"students_export_{start_date}_{end_date}.xlsx"

        else:  # ads (creatives)
            ws.title = "Creatives"

            meta_token = os.getenv("META_ACCESS_TOKEN")
            ad_account_id = os.getenv("META_AD_ACCOUNT_ID")

            if not meta_token or not ad_account_id:
                return JSONResponse({"error": "META credentials не налаштовані"}, status_code=400)

            insights = meta_conn.fetch_insights(
                ad_account_id=ad_account_id,
                access_token=meta_token,
                date_from=start_date,
                date_to=end_date,
                level="ad"
            )

            headers = ["date_start", "date_stop", "campaign_id", "campaign_name",
                       "adset_id", "adset_name", "ad_id", "ad_name",
                       "impressions", "clicks", "spend", "cpc", "cpm", "ctr"]
            ws.append(headers)

            for insight in insights:
                row = [
                    insight.get("date_start", ""),
                    insight.get("date_stop", ""),
                    insight.get("campaign_id", ""),
                    insight.get("campaign_name", ""),
                    insight.get("adset_id", ""),
                    insight.get("adset_name", ""),
                    insight.get("ad_id", ""),
                    insight.get("ad_name", ""),
                    insight.get("impressions", 0),
                    insight.get("clicks", 0),
                    insight.get("spend", 0),
                    insight.get("cpc", 0),
                    insight.get("cpm", 0),
                    insight.get("ctr", 0),
                ]
                ws.append(row)

            # Добавляем графики для рекламы
            _add_ads_charts(ws)

            filename = f"ads_export_{start_date}_{end_date}.xlsx"

        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        return FileResponse(
            temp_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


async def run_pipeline(job_id: str, params: Dict[str, Any]):
    # Create database record for this run
    with get_db() as db:
        db_run = PipelineRun(
            job_id=job_id,
            start_date=params["start_date"],
            end_date=params["end_date"],
            sheet_id=params.get("sheet_id"),
            storage_backend=os.getenv("STORAGE_BACKEND", "sheets"),
            status="running"
        )
        db.add(db_run)
        db.commit()
        db.refresh(db_run)
        run_id = db_run.id

    def log_to_db(message: str, level: str = "info"):
        """Helper to save log messages to database."""
        try:
            with get_db() as db:
                log_entry = RunLog(run_id=run_id, message=message, level=level)
                db.add(log_entry)
                db.commit()
        except Exception as e:
            print(f"Помилка запису логу в БД: {e}")

    try:
        progress.update(job_id, 2, "Перевірка облікових даних")
        log_to_db("Початок виконання pipeline")

        # Meta credentials
        meta_token = os.getenv("META_ACCESS_TOKEN")
        ad_account_id = os.getenv("META_AD_ACCOUNT_ID")
        if not meta_token or not ad_account_id:
            raise RuntimeError("META_ACCESS_TOKEN або META_AD_ACCOUNT_ID не налаштовані")

        backend = os.getenv("STORAGE_BACKEND", "sheets").lower()
        gs_client = None
        sheet_id = params.get("sheet_id")
        if backend == "sheets":
            gs_client = gs_conn.get_client()
            if not sheet_id:
                raise RuntimeError("GOOGLE_SHEET_ID обов'язковий коли STORAGE_BACKEND=sheets")

        # 1) Fetch Ads insights
        progress.update(job_id, 10, "Отримання статистики Meta Ads (рівень оголошень)")
        insights = meta_conn.fetch_insights(
            ad_account_id=ad_account_id,
            access_token=meta_token,
            date_from=params["start_date"],
            date_to=params["end_date"],
            level="ad",
        )

        # 1.5) Fetch creative details (texts and images)
        progress.update(job_id, 20, "Завантаження креативів та текстів оголошень")
        ad_ids = [insight.get("ad_id") for insight in insights if insight.get("ad_id")]
        logger.info(f"Fetching creatives for {len(ad_ids)} ads")
        creatives = meta_conn.fetch_ad_creatives(ad_ids, meta_token)

        # Merge creatives with insights
        for insight in insights:
            ad_id = insight.get("ad_id")
            if ad_id and ad_id in creatives:
                creative_data = creatives[ad_id]
                insight["creative_title"] = creative_data.get("title", "")
                insight["creative_body"] = creative_data.get("body", "")
                insight["image_url"] = creative_data.get("image_url", "")
                insight["video_id"] = creative_data.get("video_id", "")

        progress.log(job_id, f"Завантажено {len(creatives)} креативів")

        # 2) Fetch Leads
        progress.update(job_id, 30, "Підготовка даних CRM (студенти: AlfaCRM, викладачі: NetHunt)")
        # Fetch teachers from NetHunt
        teachers: list[dict] = []
        nh_folder = os.getenv("NETHUNT_FOLDER_ID")
        if nh_folder:
            try:
                raw_records = crm_tools.nethunt_list_records(nh_folder, limit=1000)
                # Flatten NetHunt records: keep id, createdAt, updatedAt and fields
                for r in raw_records:
                    flat = {}
                    if isinstance(r, dict):
                        flat["id"] = r.get("id") or r.get("_id")
                        flat["createdAt"] = r.get("createdAt")
                        flat["updatedAt"] = r.get("updatedAt")
                        fields = r.get("fields") or {}
                        if isinstance(fields, dict):
                            for k, v in fields.items():
                                # Basic normalization
                                if isinstance(v, (dict, list)):
                                    try:
                                        import json as _json
                                        flat[k] = _json.dumps(v, ensure_ascii=False)
                                    except Exception:
                                        flat[k] = str(v)
                                else:
                                    flat[k] = v
                    teachers.append(flat)
                progress.log(job_id, f"Отримано викладачів з NetHunt: {len(teachers)}")

                # Enrich teachers with funnel statistics from NetHunt journey tracking
                try:
                    progress.update(job_id, 35, "Обробка воронки статусів викладачів NetHunt")

                    # Get Meta leads grouped by campaigns for teachers
                    meta_page_id = os.getenv("META_PAGE_ID")
                    meta_page_token = os.getenv("META_PAGE_ACCESS_TOKEN")

                    if meta_page_id and meta_page_token:
                        # Фільтруємо лідів тільки для кампаній вчителів
                        keywords_teachers = os.getenv("CAMPAIGN_KEYWORDS_TEACHERS", "teacher,vchitel").lower().split(",")

                        all_campaigns = await meta_leads.get_leads_for_period(
                            page_id=meta_page_id,
                            page_token=meta_page_token,
                            start_date=params["start_date"],
                            end_date=params["end_date"]
                        )

                        # Фільтруємо тільки кампанії вчителів
                        teacher_campaigns = {
                            cid: cdata for cid, cdata in all_campaigns.items()
                            if any(kw.strip() in cdata.get("campaign_name", "").lower() for kw in keywords_teachers)
                        }

                        progress.log(job_id, f"Отримано {len(teacher_campaigns)} кампаній вчителів з {sum(len(c['leads']) for c in teacher_campaigns.values())} лідів")

                        # Трекінг через NetHunt з реальною історією
                        enriched_campaigns = await nethunt_tracking.track_leads_by_campaigns(
                            campaigns_data=teacher_campaigns,
                            folder_id=nh_folder
                        )
                        progress.log(job_id, f"Обраховано воронку для {len(enriched_campaigns)} кампаній викладачів")
                    else:
                        progress.log(job_id, "META_PAGE_ID або META_PAGE_ACCESS_TOKEN не налаштовані - пропускаємо трекінг викладачів")

                except Exception as e:
                    progress.log(job_id, f"Попередження: не вдалося обрахувати воронку викладачів: {e}")
            except Exception as e:
                progress.log(job_id, f"Помилка отримання з NetHunt: {e}")
        else:
            progress.log(job_id, "NETHUNT_FOLDER_ID не встановлено; викладачі пропущені")

        # Fetch students from AlfaCRM
        students: list[dict] = []
        if os.getenv("ALFACRM_BASE_URL") and os.getenv("ALFACRM_EMAIL") and os.getenv("ALFACRM_API_KEY") and os.getenv("ALFACRM_COMPANY_ID"):
            try:
                page = 1
                total = 0
                while True:
                    data = crm_tools.alfacrm_list_students(page=page, page_size=200)
                    # Expect data like {items:[...], total: N} or similar
                    items = data.get("items") or data.get("data") or data.get("list") or []
                    if not isinstance(items, list):
                        break
                    for s in items:
                        # Flatten student
                        flat = {}
                        if isinstance(s, dict):
                            for k, v in s.items():
                                if isinstance(v, (dict, list)):
                                    try:
                                        import json as _json
                                        flat[k] = _json.dumps(v, ensure_ascii=False)
                                    except Exception:
                                        flat[k] = str(v)
                                else:
                                    flat[k] = v
                        students.append(flat)
                    total += len(items)
                    if len(items) < 200:
                        break
                    page += 1
                progress.log(job_id, f"Отримано студентів з AlfaCRM: {len(students)}")

                # Enrich students with funnel statistics from AlfaCRM tracking
                try:
                    progress.update(job_id, 40, "Обробка воронки статусів студентів AlfaCRM")

                    # Get Meta leads grouped by campaigns for students
                    meta_page_id = os.getenv("META_PAGE_ID")
                    meta_page_token = os.getenv("META_PAGE_ACCESS_TOKEN")

                    if meta_page_id and meta_page_token:
                        # Фільтруємо лідів тільки для кампаній студентів
                        keywords_students = os.getenv("CAMPAIGN_KEYWORDS_STUDENTS", "student,shkolnik").lower().split(",")

                        all_campaigns = await meta_leads.get_leads_for_period(
                            page_id=meta_page_id,
                            page_token=meta_page_token,
                            start_date=params["start_date"],
                            end_date=params["end_date"]
                        )

                        # Фільтруємо тільки кампанії студентів
                        student_campaigns = {
                            cid: cdata for cid, cdata in all_campaigns.items()
                            if any(kw.strip() in cdata.get("campaign_name", "").lower() for kw in keywords_students)
                        }

                        progress.log(job_id, f"Отримано {len(student_campaigns)} кампаній студентів з {sum(len(c['leads']) for c in student_campaigns.values())} лідів")

                        # Трекінг через AlfaCRM з inference підходом
                        enriched_campaigns = await alfacrm_tracking.track_leads_by_campaigns(
                            campaigns_data=student_campaigns,
                            page_size=500
                        )
                        progress.log(job_id, f"Обраховано воронку для {len(enriched_campaigns)} кампаній студентів")
                    else:
                        progress.log(job_id, "META_PAGE_ID або META_PAGE_ACCESS_TOKEN не налаштовані - пропускаємо трекінг студентів")

                except Exception as e:
                    progress.log(job_id, f"Попередження: не вдалося обрахувати воронку студентів: {e}")
            except Exception as e:
                progress.log(job_id, f"Помилка отримання з AlfaCRM: {e}")
        else:
            progress.log(job_id, "AlfaCRM credentials неповні; студенти пропущені")

        # 3) Check CRM statuses
        progress.update(job_id, 45, "Отримано дані CRM")

        # 4) Write to Google Sheets
        progress.update(job_id, 70, "Запис даних у цільове сховище")
        mapping = load_mapping()
        if backend == "sheets" and gs_client:
            gs_conn.write_insights(gs_client, sheet_id, insights)
            # If needed, write students/teachers to different tabs/sheets in Sheets (not configured yet)
        else:
            excel_creatives = os.getenv("EXCEL_CREATIVES_PATH")
            excel_students = os.getenv("EXCEL_STUDENTS_PATH")
            excel_teachers = os.getenv("EXCEL_TEACHERS_PATH")
            if excel_creatives:
                cr_map = mapping.get("creatives", {})
                excel_conn.write_creatives(
                    excel_creatives,
                    insights,
                    mapping=cr_map.get("fields"),
                    sheet_name=cr_map.get("sheet_name", "Creatives"),
                )
                progress.log(job_id, f"Записано креативів: {len(insights)} рядків")
            else:
                progress.log(job_id, "EXCEL_CREATIVES_PATH не встановлено; креативи пропущені")
            if excel_students:
                st_map = mapping.get("students", {})
                excel_conn.write_students(
                    excel_students,
                    students,
                    mapping=st_map.get("fields"),
                    sheet_name=st_map.get("sheet_name", "Students"),
                )
                progress.log(job_id, f"Записано студентів: {len(students)} рядків")
            if excel_teachers:
                tc_map = mapping.get("teachers", {})
                excel_conn.write_teachers(
                    excel_teachers,
                    teachers,
                    mapping=tc_map.get("fields"),
                    sheet_name=tc_map.get("sheet_name", "Teachers"),
                )
                progress.log(job_id, f"Записано викладачів: {len(teachers)} рядків")

        progress.update(job_id, 100, "done")
        log_to_db("Pipeline завершено успішно")

        # Update database record with results
        with get_db() as db:
            db_run = db.query(PipelineRun).filter(PipelineRun.id == run_id).first()
            if db_run:
                db_run.status = "success"
                db_run.end_time = datetime.utcnow()
                db_run.insights_count = len(insights)
                db_run.students_count = len(students)
                db_run.teachers_count = len(teachers)
                db.commit()

    except Exception as e:
        progress.log(job_id, f"ERROR: {e}")
        progress.set_status(job_id, "error")
        log_to_db(f"ERROR: {e}", level="error")

        # Update database record with error
        with get_db() as db:
            db_run = db.query(PipelineRun).filter(PipelineRun.id == run_id).first()
            if db_run:
                db_run.status = "error"
                db_run.end_time = datetime.utcnow()
                db_run.error_message = str(e)
                db.commit()


@app.get("/api/students-with-journey")
@limiter.limit("30/minute")
async def get_students_with_journey(
    request: Request,
    start_date: str = None,
    end_date: str = None
):
    """
    Get students data enriched with lead journey information based on AlfaCRM status.

    Восстанавливает историю лида через воронку 38 статусов AlfaCRM
    на основе текущего lead_status_id студента.

    Query params:
    - start_date: Filter by date (YYYY-MM-DD)
    - end_date: Filter by date (YYYY-MM-DD)

    Returns:
        {
            "students": [
                {
                    ...student_data,
                    "journey_status_ids": [13, 11, 1, 32, 12, 6, 2, 3, 9, 4],
                    "journey_status_names": ["Не розібраний", "Недодзвон", ...],
                    "journey_stats": {
                        "total_steps": 10,
                        "conversion_reached": true,
                        "current_status": 4,
                        "current_status_name": "Отримана оплата",
                        "funnel_type": "main"
                    }
                }
            ],
            "count": 150
        }
    """
    try:
        from .services.lead_journey_recovery import batch_enrich_students

        # Получаем студентов из AlfaCRM
        if not (os.getenv("ALFACRM_BASE_URL") and os.getenv("ALFACRM_COMPANY_ID")):
            return JSONResponse({"error": "AlfaCRM credentials не налаштовані"}, status_code=400)

        logger.info("Fetching students from AlfaCRM with lead_status_id")

        all_students = []
        page = 1

        while True:
            data = crm_tools.alfacrm_list_students(page=page, page_size=200)
            items = data.get("items", [])

            if not items:
                break

            all_students.extend(items)

            # Проверяем есть ли еще страницы
            total = data.get("total", 0)
            if len(all_students) >= total:
                break

            page += 1

        logger.info(f"Loaded {len(all_students)} students from AlfaCRM")

        # Обогащаем студентов информацией о пути через воронку
        enriched_students = batch_enrich_students(all_students)

        # Фильтруем по датам если указаны
        if start_date or end_date:
            # Простая фильтрация по created_at или updated_at полям
            # (AlfaCRM возвращает эти поля)
            filtered = []
            for s in enriched_students:
                created = s.get("created_at") or s.get("date_create") or ""
                if start_date and str(created) < start_date:
                    continue
                if end_date and str(created) > end_date:
                    continue
                filtered.append(s)
            enriched_students = filtered

        logger.info(f"Returning {len(enriched_students)} enriched students with journey data")

        return {
            "students": enriched_students,
            "count": len(enriched_students),
            "enrichment_info": {
                "total_statuses_tracked": 38,
                "funnels": ["main", "secondary"]
            }
        }

    except Exception as e:
        logger.error(f"Error fetching students with journey: {e}")
        return JSONResponse({"error": f"Помилка отримання даних: {str(e)}"}, status_code=500)


# Serve index.html for root path
@app.get("/", response_class=HTMLResponse)
async def index():
    dist_index = os.path.join(WEB_DIST, "index.html")
    if os.path.exists(dist_index):
        with open(dist_index, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())
    static_index = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(static_index):
        with open(static_index, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())
    return HTMLResponse("<h1>UI not found</h1>")


# Mount static files AFTER all API routes to avoid conflicts
if os.path.exists(WEB_DIST):
    app.mount("/assets", StaticFiles(directory=os.path.join(WEB_DIST, "assets")), name="assets")
