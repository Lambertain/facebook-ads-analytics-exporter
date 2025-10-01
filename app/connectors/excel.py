import os
import logging
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _ensure_book(path: str):
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            logger.debug(f"Loaded existing Excel workbook: {path}")
            return wb
        except Exception as e:
            logger.warning(f"Excel file corrupted or locked ({type(e).__name__}), recreating: {path}")
            wb = Workbook()
            wb.save(path)
            return wb
    logger.info(f"Creating new Excel workbook: {path}")
    wb = Workbook()
    wb.save(path)
    return wb


def _write_table(path: str, sheet_name: str, headers: List[str], rows: List[List[Any]]):
    """Create/overwrite a sheet completely (legacy helper)."""
    try:
        wb = _ensure_book(path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            logger.debug(f"Deleted existing sheet '{sheet_name}'")
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for r in rows:
            ws.append(r)
        for idx, h in enumerate(headers, start=1):
            max_len = max([len(str(h))] + [len(str(row[idx-1])) if idx-1 < len(row) and row[idx-1] is not None else 0 for row in rows])
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max(10, max_len + 2))
        wb.save(path)
        logger.info(f"Successfully wrote {len(rows)} rows to sheet '{sheet_name}' in {path}")
    except Exception as e:
        logger.error(f"Failed to write table to {path} sheet '{sheet_name}': {type(e).__name__}: {e}")
        raise


def _write_by_headers(path: str, sheet_name: str, rows_as_dicts: List[Dict[str, Any]]):
    """Write rows aligning to existing header row in the given sheet.

    - If sheet doesn't exist, creates it with headers based on the first row dict keys.
    - Keeps headers as-is if sheet exists; clears data from row 2 down and writes fresh content.
    """
    try:
        wb = _ensure_book(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
            headers = [str(c) if c is not None else '' for c in header_row]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
                logger.debug(f"Cleared {ws.max_row - 1} existing data rows from sheet '{sheet_name}'")
        else:
            ws = wb.create_sheet(title=sheet_name)
            headers = list(rows_as_dicts[0].keys()) if rows_as_dicts else []
            if headers:
                ws.append(headers)
            logger.info(f"Created new sheet '{sheet_name}' with {len(headers)} headers")

        for d in rows_as_dicts:
            row = [d.get(h) for h in headers]
            ws.append(row)
        wb.save(path)
        logger.info(f"Successfully wrote {len(rows_as_dicts)} rows by headers to sheet '{sheet_name}' in {path}")
    except Exception as e:
        logger.error(f"Failed to write by headers to {path} sheet '{sheet_name}': {type(e).__name__}: {e}")
        raise


def write_creatives(path: str, insights: List[Dict[str, Any]], mapping: Dict[str, str] = None, sheet_name: str = "Creatives"):
    # mapping: api_field -> header_name
    rows = []
    for i in insights:
        if mapping:
            rows.append({ header: i.get(api_field) for api_field, header in mapping.items() })
        else:
            rows.append(i)
    _write_by_headers(path, sheet_name=sheet_name, rows_as_dicts=rows)


def write_students(path: str, rows_dicts: List[Dict[str, Any]], mapping: Dict[str, str] = None, sheet_name: str = "Students"):
    if mapping:
        rows = [{ header: d.get(api_field) for api_field, header in mapping.items() } for d in rows_dicts]
    else:
        rows = rows_dicts
    _write_by_headers(path, sheet_name=sheet_name, rows_as_dicts=rows)


def write_teachers(path: str, rows_dicts: List[Dict[str, Any]], mapping: Dict[str, str] = None, sheet_name: str = "Teachers"):
    if mapping:
        rows = [{ header: d.get(api_field) for api_field, header in mapping.items() } for d in rows_dicts]
    else:
        rows = rows_dicts
    _write_by_headers(path, sheet_name=sheet_name, rows_as_dicts=rows)
