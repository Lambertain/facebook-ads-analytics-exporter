#!/usr/bin/env python3
"""Analyze Excel file from production using openpyxl."""
from openpyxl import load_workbook
import sys

try:
    excel_file = 'ecademy_meta_data_2025-10-23.xlsx'
    wb = load_workbook(excel_file, read_only=True, data_only=True)

    print("=" * 80)
    print("SHEETS IN FILE:")
    print("=" * 80)
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")

    print("\n" + "=" * 80)

    # Check STUDENTS sheet
    if 'Студенти' in wb.sheetnames:
        ws = wb['Студенти']
        print("STUDENTS SHEET (Студенти):")
        print("=" * 80)
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")

        # Get headers (first row)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)

        print(f"\nTotal columns: {len(headers)}")
        print("\nCOLUMNS:")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")

        # Show first 3 data rows
        print("\n" + "=" * 80)
        print("FIRST 3 DATA ROWS:")
        print("=" * 80)
        for row_idx in range(2, min(5, ws.max_row + 1)):  # rows 2-4 (skip header)
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):  # First 10 columns
                cell_value = ws.cell(row=row_idx, column=col).value
                row_data.append(str(cell_value)[:30] if cell_value else "")
            print(f"Row {row_idx-1}: {' | '.join(row_data)}")

        # Count zero values in numeric columns (sample first 20 columns)
        print("\n" + "=" * 80)
        print("COLUMNS WITH MANY ZEROS (first 20 cols):")
        print("=" * 80)
        for col_idx in range(1, min(21, ws.max_column + 1)):
            header = headers[col_idx - 1]
            zero_count = 0
            total_count = ws.max_row - 1  # excluding header

            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if value == 0 or value == "0":
                    zero_count += 1

            if total_count > 0:
                zero_percent = (zero_count / total_count * 100)
                if zero_percent > 50:
                    print(f"  Col {col_idx} ({header}): {zero_percent:.1f}% zeros ({zero_count}/{total_count})")

    else:
        print("ERROR: Sheet 'Студенти' not found!")

    wb.close()

except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
