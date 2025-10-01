import sys
from pathlib import Path
from openpyxl import load_workbook


def print_headers(xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"FILE: {xlsx_path}")
    for name in wb.sheetnames:
        ws = wb[name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(c) if c is not None else '' for c in (row1 or [])]
        print(f"  SHEET: {name}")
        print(f"    HEADERS: {headers}")
    print()


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/inspect_excel.py <xlsx> [<xlsx> ...]")
        sys.exit(1)
    for p in sys.argv[1:]:
        print_headers(Path(p))


if __name__ == "__main__":
    main()

