import json
import os
import re
from pathlib import Path
from typing import Dict, List

from dotenv import load_dotenv
from openpyxl import load_workbook


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[\s_\-]+", " ", s)
    s = s.replace("ё", "е")
    return s


def read_headers(path: str) -> Dict[str, List[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        headers = [str(c).strip() if c is not None else '' for c in row1]
        # drop trailing empties
        while headers and headers[-1] == '':
            headers.pop()
        sheets[name] = headers
    return sheets


def choose_sheet_name(sheet_dict: Dict[str, List[str]], preferred: str) -> str:
    if preferred in sheet_dict:
        return preferred
    if sheet_dict:
        return list(sheet_dict.keys())[0]
    return preferred


def generate_creatives_mapping(headers: List[str]) -> Dict[str, str]:
    # source field -> header name
    known = {
        'date_start': {'date start', 'дата начала', 'дата'},
        'date_stop': {'date stop', 'дата окончания', 'до'},
        'campaign_id': {'campaign id', 'id кампании', 'id кампании'},
        'campaign_name': {'campaign', 'название кампании', 'кампания'},
        'adset_id': {'adset id', 'id набора объявлений'},
        'adset_name': {'adset', 'набор объявлений'},
        'ad_id': {'ad id', 'id объявления', 'id креатива'},
        'ad_name': {'ad name', 'объявление', 'креатив', 'название объявления'},
        'impressions': {'impressions', 'показы'},
        'clicks': {'clicks', 'клики'},
        'spend': {'spend', 'расход', 'затраты', 'стоимость'},
        'cpc': {'cpc', 'цена клика'},
        'cpm': {'cpm'},
        'ctr': {'ctr'},
    }
    norm_headers = {h: norm(h) for h in headers}
    mapping: Dict[str, str] = {}
    for key, synonyms in known.items():
        target = None
        for h, nh in norm_headers.items():
            if nh == key or nh in synonyms:
                target = h
                break
        mapping[key] = target or key
    return mapping


def generate_passthrough_mapping(headers: List[str], extras: List[str] = None) -> Dict[str, str]:
    mapping = {}
    for h in headers:
        if h:
            mapping[h] = h
    for e in extras or []:
        # map known extras to similar header if exists
        cand = next((h for h in headers if norm(h) == norm(e)), None)
        mapping[e] = cand or e
    return mapping


def main():
    root = Path(__file__).resolve().parents[1]
    load_dotenv(root / '.env')
    creatives = os.getenv('EXCEL_CREATIVES_PATH')
    students = os.getenv('EXCEL_STUDENTS_PATH')
    teachers = os.getenv('EXCEL_TEACHERS_PATH')
    out_path = root / 'config' / 'mapping.json'
    out_path.parent.mkdir(parents=True, exist_ok=True)

    mapping = { 'creatives': {}, 'students': {}, 'teachers': {} }

    if creatives and os.path.exists(creatives):
        sheets = read_headers(creatives)
        sheet_name = choose_sheet_name(sheets, 'Creatives')
        mapping['creatives'] = {
            'sheet_name': sheet_name,
            'fields': generate_creatives_mapping(sheets.get(sheet_name, []))
        }
    if students and os.path.exists(students):
        sheets = read_headers(students)
        sheet_name = choose_sheet_name(sheets, 'Students')
        mapping['students'] = {
            'sheet_name': sheet_name,
            'fields': generate_passthrough_mapping(sheets.get(sheet_name, []))
        }
    if teachers and os.path.exists(teachers):
        sheets = read_headers(teachers)
        sheet_name = choose_sheet_name(sheets, 'Teachers')
        mapping['teachers'] = {
            'sheet_name': sheet_name,
            'fields': generate_passthrough_mapping(sheets.get(sheet_name, []), extras=['id','createdAt','updatedAt'])
        }

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)
    print(f'Wrote draft mapping to {out_path}')


if __name__ == '__main__':
    main()

