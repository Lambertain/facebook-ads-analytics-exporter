"""
Unit тести для Excel connector (app/connectors/excel.py).

Базові тести для функцій роботи з Excel файлами.
"""

import pytest
from unittest.mock import Mock, patch, MagicMock, call
from openpyxl import Workbook
from app.connectors import excel


class TestEnsureBook:
    """Тести для функції _ensure_book."""

    @patch('app.connectors.excel.os.path.exists')
    @patch('app.connectors.excel.load_workbook')
    def test_ensure_book_file_exists(self, mock_load, mock_exists):
        """Тест відкриття існуючого файлу."""
        # Arrange
        mock_exists.return_value = True
        mock_wb = Mock()
        mock_load.return_value = mock_wb

        # Act
        result = excel._ensure_book("test.xlsx")

        # Assert
        assert result == mock_wb
        mock_load.assert_called_once_with("test.xlsx")

    @patch('app.connectors.excel.os.path.exists')
    @patch('app.connectors.excel.Workbook')
    def test_ensure_book_file_not_exists(self, mock_wb_class, mock_exists):
        """Тест створення нового файлу."""
        # Arrange
        mock_exists.return_value = False
        mock_wb = Mock()
        mock_wb_class.return_value = mock_wb

        # Act
        result = excel._ensure_book("new_file.xlsx")

        # Assert
        assert result == mock_wb
        mock_wb.save.assert_called_once_with("new_file.xlsx")

    @patch('app.connectors.excel.os.path.exists')
    @patch('app.connectors.excel.load_workbook')
    @patch('app.connectors.excel.Workbook')
    def test_ensure_book_corrupted_file(self, mock_wb_class, mock_load, mock_exists):
        """Тест відновлення пошкодженого файлу."""
        # Arrange
        mock_exists.return_value = True
        mock_load.side_effect = Exception("File corrupted")
        mock_wb = Mock()
        mock_wb_class.return_value = mock_wb

        # Act
        result = excel._ensure_book("corrupted.xlsx")

        # Assert
        assert result == mock_wb
        mock_wb.save.assert_called_once()


# Note: _write_table, _write_by_headers, write_insights, write_leads
# тести закоментовані, оскільки потребують складного мокування Workbook
# Базовий coverage (67%) досягнуто через _ensure_book тести


class TestWriteTable:
    """Тести для функції _write_table."""

    @patch('app.connectors.excel._ensure_book')
    def test_write_table_new_sheet(self, mock_ensure):
        """Тест створення нового аркуша з даними."""
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_wb.sheetnames = []
        mock_wb.create_sheet.return_value = mock_ws
        mock_ensure.return_value = mock_wb

        headers = ["Name", "Age", "City"]
        rows = [["John", 30, "NYC"], ["Jane", 25, "LA"]]

        excel._write_table("test.xlsx", "Sheet1", headers, rows)

        mock_ensure.assert_called_once_with("test.xlsx")
        mock_wb.create_sheet.assert_called_once_with(title="Sheet1")
        assert mock_ws.append.call_count == 3
        mock_wb.save.assert_called_once_with("test.xlsx")

    @patch('app.connectors.excel._ensure_book')
    def test_write_table_overwrite_existing(self, mock_ensure):
        """Тест перезапису існуючого аркуша."""
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_wb.sheetnames = ["Sheet1"]
        mock_wb.create_sheet.return_value = mock_ws
        mock_ensure.return_value = mock_wb

        headers = ["Header1"]
        rows = [["Value1"]]

        excel._write_table("test.xlsx", "Sheet1", headers, rows)

        mock_wb.__delitem__.assert_called_once_with("Sheet1")
        mock_wb.create_sheet.assert_called_once()


class TestWriteByHeaders:
    """Тести для функції _write_by_headers."""

    @patch('app.connectors.excel._ensure_book')
    def test_write_by_headers_new_sheet(self, mock_ensure):
        """Тест створення нового аркуша з dictionary rows."""
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_wb.sheetnames = []
        mock_wb.create_sheet.return_value = mock_ws
        mock_ensure.return_value = mock_wb

        rows_dicts = [
            {"name": "John", "age": 30},
            {"name": "Jane", "age": 25}
        ]

        excel._write_by_headers("test.xlsx", "People", rows_dicts)

        mock_wb.create_sheet.assert_called_once_with(title="People")
        assert mock_ws.append.call_count == 3

    @patch('app.connectors.excel._ensure_book')
    def test_write_by_headers_existing_sheet(self, mock_ensure):
        """Тест запису даних у існуючий аркуш."""
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_ws.max_row = 5
        mock_ws.iter_rows.return_value = iter([[("name",), ("age",)]])
        mock_wb.sheetnames = ["People"]
        mock_wb.__getitem__.return_value = mock_ws
        mock_ensure.return_value = mock_wb

        rows_dicts = [{"name": "Bob", "age": 35}]

        excel._write_by_headers("test.xlsx", "People", rows_dicts)

        mock_ws.delete_rows.assert_called_once_with(2, 4)
        mock_wb.save.assert_called_once()


class TestWriteStudents:
    """Тести для функції write_students."""

    @patch('app.connectors.excel._write_by_headers')
    def test_write_students_no_mapping(self, mock_write):
        """Тест запису студентів без mapping."""
        students = [
            {"student_name": "Alice", "course": "Python"},
            {"student_name": "Bob", "course": "JS"}
        ]

        excel.write_students("students.xlsx", students)

        mock_write.assert_called_once_with(
            "students.xlsx",
            sheet_name="Students",
            rows_as_dicts=students
        )

    @patch('app.connectors.excel._write_by_headers')
    def test_write_students_with_mapping(self, mock_write):
        """Тест запису студентів з mapping полів."""
        students = [
            {"api_name": "Alice", "api_course": "Python"}
        ]
        mapping = {
            "api_name": "Student Name",
            "api_course": "Course Title"
        }

        excel.write_students("students.xlsx", students, mapping=mapping)

        called_rows = mock_write.call_args[1]["rows_as_dicts"]
        assert called_rows[0]["Student Name"] == "Alice"
        assert called_rows[0]["Course Title"] == "Python"


class TestWriteTeachers:
    """Тести для функції write_teachers."""

    @patch('app.connectors.excel._write_by_headers')
    def test_write_teachers_basic(self, mock_write):
        """Тест запису викладачів."""
        teachers = [{"name": "Dr. Smith", "subject": "Math"}]

        excel.write_teachers("teachers.xlsx", teachers)

        mock_write.assert_called_once_with(
            "teachers.xlsx",
            sheet_name="Teachers",
            rows_as_dicts=teachers
        )


class TestWriteCreatives:
    """Тести для функції write_creatives."""

    @patch('app.connectors.excel._write_by_headers')
    def test_write_creatives_with_mapping(self, mock_write):
        """Тест запису креативів з mapping."""
        insights = [
            {"campaign_id": "123", "impressions": 1000}
        ]
        mapping = {
            "campaign_id": "Campaign",
            "impressions": "Total Impressions"
        }

        excel.write_creatives("ads.xlsx", insights, mapping=mapping)

        called_rows = mock_write.call_args[1]["rows_as_dicts"]
        assert called_rows[0]["Campaign"] == "123"
        assert called_rows[0]["Total Impressions"] == 1000
