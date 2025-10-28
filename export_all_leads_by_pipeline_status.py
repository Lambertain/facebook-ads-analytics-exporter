"""
Експорт ВСІХ лідів з AlfaCRM в Excel таблицю по статусах воронки.

Формат:
- Стовпці = Статуси воронки (назва на укр)
- Кожна ячейка = 1 лід з повною інформацією
- Інформація в ячейці = багаторядковий текст "Назва (укр) - значення"
"""
import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
import requests
from collections import defaultdict

# Загрузка переменных окружения
load_dotenv()

# Добавить app директорию в Python path
app_dir = Path(__file__).parent / "app"
sys.path.insert(0, str(app_dir))

from connectors.crm import alfacrm_auth_get_token


# Словарь переводов названий полей на украинский
FIELD_TRANSLATIONS = {
    "id": "ID",
    "name": "Ім'я",
    "created_at": "Дата створення",
    "updated_at": "Дата оновлення",
    "lead_status_id": "ID статусу ліда",
    "study_status_id": "ID статусу навчання",
    "is_study": "Навчається",
    "color": "Колір",
    "phone": "Телефони",
    "email": "Email",
    "addr": "Адреса",
    "balance": "Баланс",
    "balance_base": "Базовий баланс",
    "balance_bonus": "Бонусний баланс",
    "paid_count": "Кількість оплат",
    "paid_lesson_count": "Кількість оплачених уроків",
    "assigned_id": "ID відповідального",
    "branch_ids": "ID філій",
    "company_id": "ID компанії",
    "b_date": "Дата народження",
    "sex": "Стать",
    "barcode": "Штрих-код",
    "comment": "Коментар",
    "legal_type": "Тип особи",
    "pipeline_id": "ID воронки",
    "custom_ads_comp": "Назва кампанії",
    "custom_id_srm": "ID з Facebook",
    "custom_gorodstvaniya": "Громадянство",
    "custom_age_": "Вік",
    "custom_email": "Email (додатковий)",
    "custom_yazik": "Мова",
    "custom_urovenvladenwoo": "Рівень володіння",
    "custom_schedule": "Розклад",
    "custom_try_lessons": "Пробні уроки",
}


def get_pipelines_and_statuses():
    """
    Получить список воронок и их статусов из AlfaCRM.

    Returns:
        dict: {pipeline_id: {"name": "...", "statuses": {status_id: "name"}}}
    """
    print("\n[1/4] Отримання воронок та статусів...")

    token = alfacrm_auth_get_token()
    base_url = os.getenv("ALFACRM_BASE_URL").rstrip('/')
    company_id = int(os.getenv("ALFACRM_COMPANY_ID"))

    # Получаем список статусов лидов
    try:
        url = f"{base_url}/v2api/lead-status/index"
        payload = {"branch_id": company_id}

        resp = requests.post(
            url,
            headers={"X-ALFACRM-TOKEN": token},
            json=payload,
            timeout=15
        )
        resp.raise_for_status()
        data = resp.json()

        statuses = data.get("items", [])
        print(f"  ✓ Отримано {len(statuses)} статусів лідів")

        # Группируем по воронкам
        pipelines = defaultdict(lambda: {"name": "", "statuses": {}})

        for status in statuses:
            status_id = status.get("id")
            status_name = status.get("name", f"Статус {status_id}")
            pipeline_id = status.get("pipeline_id", 0)

            if pipeline_id not in pipelines:
                pipelines[pipeline_id] = {
                    "name": f"Воронка {pipeline_id}",
                    "statuses": {}
                }

            pipelines[pipeline_id]["statuses"][status_id] = status_name

        print(f"  ✓ Знайдено {len(pipelines)} воронок")
        for pid, pdata in pipelines.items():
            print(f"    - {pdata['name']}: {len(pdata['statuses'])} статусів")

        return dict(pipelines)

    except Exception as e:
        print(f"  ⚠ Не вдалося отримати статуси через API: {e}")
        print(f"  → Використовую базові статуси")

        # Fallback: базові статуси на основі того что видели в данных
        return {
            0: {
                "name": "Основна воронка",
                "statuses": {
                    4: "Активний",
                    39: "Архів",
                }
            }
        }


def get_all_leads():
    """
    Получить ВСЕХ лидов из AlfaCRM.

    Returns:
        List[dict]: Список всех лидов
    """
    print("\n[2/4] Завантаження всіх лідів...")

    token = alfacrm_auth_get_token()
    base_url = os.getenv("ALFACRM_BASE_URL").rstrip('/')
    company_id = int(os.getenv("ALFACRM_COMPANY_ID"))

    all_leads = []
    page = 1
    page_size = 500

    while True:
        url = f"{base_url}/v2api/customer/index"
        payload = {
            "branch_ids": [company_id],
            "page": page,
            "page_size": page_size
        }

        try:
            resp = requests.post(
                url,
                headers={"X-ALFACRM-TOKEN": token},
                json=payload,
                timeout=15
            )
            resp.raise_for_status()
            data = resp.json()

            items = data.get("items", [])
            total = data.get("count", 0)

            print(f"  Сторінка {page}: {len(items)} лідів")

            if not items:
                break

            all_leads.extend(items)

            if len(all_leads) >= total:
                break

            page += 1

        except Exception as e:
            print(f"  ✗ Помилка: {e}")
            break

    print(f"  ✓ Всього завантажено: {len(all_leads)} лідів")
    return all_leads


def format_lead_info(lead: dict) -> str:
    """
    Форматувати інформацію про ліда для відображення в ячейці.

    Args:
        lead: Словник з даними про ліда

    Returns:
        str: Багаторядковий текст "Назва - значення"
    """
    lines = []

    # Сортируем поля: важные сначала, потом остальные
    important_fields = [
        "id", "name", "phone", "email",
        "created_at", "balance", "paid_count"
    ]

    # Сначала важные поля
    for field in important_fields:
        if field in lead:
            value = lead[field]
            if value is None or value == "" or value == []:
                continue

            ukr_name = FIELD_TRANSLATIONS.get(field, field)
            formatted_value = format_value_for_cell(value)
            lines.append(f"{ukr_name}: {formatted_value}")

    # Потом остальные поля
    for field, value in sorted(lead.items()):
        if field in important_fields:
            continue

        if value is None or value == "" or value == []:
            continue

        ukr_name = FIELD_TRANSLATIONS.get(field, field.replace("_", " ").title())
        formatted_value = format_value_for_cell(value)
        lines.append(f"{ukr_name}: {formatted_value}")

    return "\n".join(lines)


def format_value_for_cell(value):
    """
    Форматировать значение для отображения в ячейке.
    """
    if value is None or value == "":
        return ""

    if isinstance(value, list):
        if not value:
            return ""
        return ", ".join(str(v) for v in value)

    if isinstance(value, bool):
        return "Так" if value else "Ні"

    if isinstance(value, dict):
        return str(value)

    # Ограничиваем длину очень длинных строк
    if isinstance(value, str) and len(value) > 200:
        return value[:200] + "..."

    return str(value)


def export_to_excel(leads: list, pipelines: dict, output_file: str):
    """
    Экспортировать лиды в Excel по статусам воронки.

    Args:
        leads: Список лидов
        pipelines: Информация о воронках и статусах
        output_file: Путь к выходному файлу
    """
    print(f"\n[3/4] Групування лідів по статусах...")

    # Группируем лиды по статусам
    leads_by_status = defaultdict(list)

    for lead in leads:
        lead_status = lead.get("lead_status_id")
        pipeline_id = lead.get("pipeline_id", 0)

        # Ключ: (pipeline_id, status_id)
        key = (pipeline_id, lead_status)
        leads_by_status[key].append(lead)

    print(f"  ✓ Лідів згруповано по {len(leads_by_status)} статусах")

    # Создаем структуру для Excel
    print(f"\n[4/4] Створення Excel таблиці...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Ліди по статусах"

        # Создаем заголовки столбцов
        columns = []
        col_index = 1

        for pipeline_id, pipeline_data in sorted(pipelines.items()):
            pipeline_name = pipeline_data["name"]
            statuses = pipeline_data["statuses"]

            for status_id, status_name in sorted(statuses.items()):
                # Название столбца
                if len(pipelines) > 1:
                    col_name = f"{status_name} - {pipeline_name}"
                else:
                    col_name = status_name

                columns.append({
                    "name": col_name,
                    "pipeline_id": pipeline_id,
                    "status_id": status_id,
                    "col_index": col_index
                })

                # Записываем заголовок
                cell = ws.cell(row=1, column=col_index)
                cell.value = col_name
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                col_index += 1

        print(f"  ✓ Створено {len(columns)} стовпців")

        # Заполняем данные
        max_rows_in_column = defaultdict(int)

        for col_data in columns:
            pipeline_id = col_data["pipeline_id"]
            status_id = col_data["status_id"]
            col_index = col_data["col_index"]

            # Получаем лиды для этого статуса
            key = (pipeline_id, status_id)
            column_leads = leads_by_status.get(key, [])

            # Записываем каждого лида в отдельную ячейку (строку)
            for row_offset, lead in enumerate(column_leads, start=2):
                cell = ws.cell(row=row_offset, column=col_index)
                cell.value = format_lead_info(lead)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Отслеживаем максимальную заполненную строку
                max_rows_in_column[col_index] = max(max_rows_in_column[col_index], row_offset)

        # Настраиваем ширину столбцов и высоту строк
        for col_data in columns:
            col_letter = ws.cell(row=1, column=col_data["col_index"]).column_letter
            ws.column_dimensions[col_letter].width = 40  # Ширина столбца

        # Устанавливаем высоту строк для данных
        for row_idx in range(2, max(max_rows_in_column.values()) + 1 if max_rows_in_column else 2):
            ws.row_dimensions[row_idx].height = 150  # Высота ячейки

        # Замораживаем первую строку
        ws.freeze_panes = "A2"

        # Сохраняем файл
        wb.save(output_file)
        print(f"  ✓ Файл збережено: {output_file}")

        # Статистика
        print(f"\n{'='*80}")
        print("СТАТИСТИКА ПО СТАТУСАХ:")
        print(f"{'='*80}\n")

        for col_data in columns:
            pipeline_id = col_data["pipeline_id"]
            status_id = col_data["status_id"]
            col_name = col_data["name"]

            key = (pipeline_id, status_id)
            count = len(leads_by_status.get(key, []))

            if count > 0:
                bar_length = min(int(count / 2), 40)
                bar = "█" * bar_length
                print(f"{col_name:40} │ {count:3} лідів │ {bar}")

        print(f"\n{'='*80}\n")

    except ImportError:
        print("  ✗ Помилка: необхідна бібліотека openpyxl")
        print("  Встановіть: pip install openpyxl")
    except Exception as e:
        print(f"  ✗ Помилка при створенні Excel: {e}")
        import traceback
        traceback.print_exc()


def main():
    """
    Основна функція.
    """
    print("\n" + "="*80)
    print("ЕКСПОРТ ВСІХ ЛІДІВ З ALFACRM ПО СТАТУСАХ ВОРОНКИ")
    print("="*80)

    # 1. Получаем воронки и статусы
    pipelines = get_pipelines_and_statuses()

    # 2. Получаем всех лидов
    leads = get_all_leads()

    if not leads:
        print("\n❌ Не вдалося отримати лідів")
        return

    # 3. Экспортируем в Excel
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"Всі_ліди_по_статусах_{timestamp}.xlsx"

    export_to_excel(leads, pipelines, output_file)

    print(f"✅ ГОТОВО! Експортовано {len(leads)} лідів")
    print(f"📄 Файл: {output_file}\n")


if __name__ == "__main__":
    main()
