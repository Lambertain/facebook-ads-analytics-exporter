"""
Скрипт для очищення даних студентів з Excel файлу.
Залишає тільки заголовки.
"""
import openpyxl
import os

excel_path = r"D:\Automation\n8n\Клиенты\eCademy\Таблицы\Students Analysis.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Видаляємо всі рядки крім заголовків (перший рядок)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    wb.save(excel_path)
    print(f"Очищено {excel_path}")
    print(f"Залишено тільки заголовки (рядок 1)")
else:
    print(f"Файл не знайдено: {excel_path}")
